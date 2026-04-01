"""
صفحة إدارة المنتجات والمخزون
Products & Inventory Management
"""

import os
import mysql.connector
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit,
                             QLabel, QTableWidget, QTableWidgetItem, QMessageBox,
                             QComboBox, QDoubleSpinBox, QSpinBox, QFileDialog,
                             QDialog, QFormLayout, QHeaderView, QTabWidget, QScrollArea, 
                             QAbstractItemView, QDateEdit, QCheckBox, QProgressDialog,
                             QApplication)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont, QColor
from PyQt6.QtPrintSupport import QPrinterInfo
from database_manager import DatabaseManager
from ui.styles import GLOBAL_STYLE, BUTTON_STYLES, get_button_style, COLORS, TABLE_STYLE, GROUP_BOX_STYLE, INPUT_STYLE, LABEL_STYLE_HEADER, LABEL_STYLE_TITLE, TAB_STYLE, CARD_STYLE
import openpyxl
from datetime import datetime
from time import perf_counter
from utils.printer_service import PrinterService

class ProductsPage(QWidget):
    """صفحة إدارة المنتجات والمخزون"""
    data_changed = pyqtSignal()
    
    def __init__(self, user_info=None, parent=None, auto_load=True):
        super().__init__(parent)
        self.setStyleSheet(GLOBAL_STYLE)
        self.db = DatabaseManager()
        self.user_info = user_info or {}
        self.auto_load = auto_load
        self.can_view_buy_price = True
        self._products_total_count = 0
        self._inventory_total_count = 0
        self._products_loaded = False
        self._inventory_loaded = False
        self._barcode_loaded = False
        self._max_barcode_rows = 300
        self.init_ui()
        if self.auto_load:
            self.load_products()
            self.load_inventory()
            self.load_barcode_products()
    
    def set_user(self, user_info):
        """تعيين بيانات المستخدم"""
        self.user_info = user_info or {}
        
        # جلب اسم المتجر الحالي
        self.current_store_name = ""
        try:
            store_id = self.user_info.get('store_id')
            if store_id:
                cursor = self.db.cursor
                cursor.execute("SELECT store_name FROM stores WHERE id = %s", (store_id,))
                res = cursor.fetchone()
                if res and 'store_name' in res:
                    self.current_store_name = res['store_name']
        except (mysql.connector.Error, AttributeError, KeyError, TypeError):
            pass

        self.check_permissions()
        if self.auto_load:
            self.current_products_page = 1
            self.current_inventory_page = 1
            self.load_products()
            self.load_inventory()
        # تحديث التبويبات الأخرى
        if hasattr(self, 'incoming_transfers_widget'):
            self.incoming_transfers_widget.set_user(self.user_info)
        if hasattr(self, 'category_mgmt_widget'):
            self.category_mgmt_widget.set_user(self.user_info)

    def ensure_loaded(self, force=False):
        """تحميل بيانات الصفحة عند الطلب (Lazy Loading)."""
        if force or not self._products_loaded:
            self.load_products()
        if force or not self._inventory_loaded:
            self.load_inventory()
        if force or not self._barcode_loaded:
            self.load_barcode_products()
        if hasattr(self, 'incoming_transfers_widget') and hasattr(self.incoming_transfers_widget, 'ensure_loaded'):
            self.incoming_transfers_widget.ensure_loaded(force=force)
        if hasattr(self, 'category_mgmt_widget') and hasattr(self.category_mgmt_widget, 'ensure_loaded'):
            self.category_mgmt_widget.ensure_loaded(force=force)

    def check_permissions(self):
        """التحقق من الصلاحيات لإظهار/إخفاء أزرار الإدارة"""
        role_name = self.user_info.get('role_name', '').lower()
        role_id = self.user_info.get('role_id')
        
        # صلاحيات تعديل المنتجات (إضافة، استيراد، تعديل)
        allowed_roles = ['admin', 'system admin', 'مدير عام', 'developer']
        is_admin = any(role in role_name for role in allowed_roles) or (role_id in [99, 3])
        
        # صلاحية رؤية سعر الشراء (حصراً لمدير النظام والمطور والكاشير بعد الترقية)
        # Role 1 is Admin, Role 99 is Developer, Role 3 is Cashier
        self.can_view_buy_price = (role_id in [1, 99, 3])
        
        if hasattr(self, 'add_btn'):
            self.add_btn.setVisible(is_admin)
        if hasattr(self, 'import_btn'):
            self.import_btn.setVisible(is_admin)
        if hasattr(self, 'fix_costs_btn'):
            self.fix_costs_btn.setVisible(is_admin and self.can_view_buy_price)
        
        # تحديث ظهور تبويب إدارة الفئات
        if hasattr(self, 'tabs') and hasattr(self, 'category_mgmt_widget'):
            idx = self.tabs.indexOf(self.category_mgmt_widget)
            if idx != -1:
                self.tabs.setTabVisible(idx, is_admin)

    def init_ui(self):
        """إنشاء واجهة الصفحة"""
        layout = QVBoxLayout()
        
        # العنوان
        title = QLabel("📦 إدارة المنتجات والمخزون")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # التبويبات
        self.tabs = QTabWidget()
        
        # تبويب المنتجات
        products_widget = self.create_products_tab()
        self.tabs.addTab(products_widget, "المنتجات")
        
        # تبويب المخزون
        inventory_widget = self.create_inventory_tab()
        self.tabs.addTab(inventory_widget, "المخزون")

        # تبويب التحويلات الواردة
        self.incoming_transfers_widget = IncomingTransfersWidget(self.db, self.user_info, auto_load=False)
        self.tabs.addTab(self.incoming_transfers_widget, "التحويلات الواردة")
        
        # تبويب إدارة الفئات (للمسؤولين فقط)
        self.category_mgmt_widget = CategoryManagementWidget(self.db, self.user_info, auto_load=False)
        self.tabs.addTab(self.category_mgmt_widget, "إدارة الفئات")

        # تبويب طباعة الباركود
        barcode_widget = self.create_barcode_tab()
        self.tabs.addTab(barcode_widget, "طباعة الباركود")
        
        layout.addWidget(self.tabs)
        
        # التحقق من الصلاحيات لإظهار/إخفاء تبويب الفئات
        self.check_permissions()
        
        # Populate supplier filters
        self.populate_suppliers_combos()
        
        self.setLayout(layout)
    
    def create_products_tab(self):
        """إنشاء تبويب المنتجات"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 1. Header Card with Product Count
        from PyQt6.QtWidgets import QFrame
        header_card = QFrame()
        header_card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['bg_main']}, stop:1 {COLORS['primary']});
                border: 2px solid {COLORS['border']};
                border-radius: 16px;
                padding: 20px;
            }}
        """)
        header_layout = QVBoxLayout(header_card)
        
        title = QLabel("📦 إدارة المنتجات والمخزون")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_light']}; border: none; background: transparent;")
        header_layout.addWidget(title)
        
        self.products_count_label = QLabel("إجمالي المنتجات: 0 منتج")
        self.products_count_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 14px; border: none; background: transparent;")
        header_layout.addWidget(self.products_count_label)
        
        layout.addWidget(header_card)
        
        # 2. Action Buttons Row
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.add_btn = QPushButton("➕ إضافة منتج جديد")
        self.add_btn.setStyleSheet(get_button_style('success'))
        self.add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.add_btn.clicked.connect(self.add_product)
        button_layout.addWidget(self.add_btn)
        
        self.import_btn = QPushButton("📥 استيراد من Excel")
        self.import_btn.setStyleSheet(get_button_style('info'))
        self.import_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.import_btn.clicked.connect(self.import_from_excel)
        button_layout.addWidget(self.import_btn)

        self.fix_costs_btn = QPushButton("🛠️ تصحيح التكاليف")
        self.fix_costs_btn.setStyleSheet(get_button_style('warning'))
        self.fix_costs_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fix_costs_btn.setToolTip("تحديث تكلفة المنتجات التي ليس لها سعر شراء (بناءً على هامش ربح افتراضي 25%)")
        self.fix_costs_btn.clicked.connect(self.fix_zero_costs_dialog)
        button_layout.addWidget(self.fix_costs_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # 3. Search Bar (Separate Row)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 ابحث عن منتج بالاسم أو الكود...")
        self.search_input.setFont(QFont("Arial", 13))
        self.search_input.setStyleSheet(INPUT_STYLE + "min-height: 50px;")
        self.search_input.textChanged.connect(self.filter_products)
        layout.addWidget(self.search_input)
        
        # --- Advanced Filters Row ---
        filters_layout = QHBoxLayout()
        filters_layout.setSpacing(10)
        
        # Supplier Filter
        filters_layout.addWidget(QLabel("المورد:"))
        self.prod_supplier_filter = QComboBox()
        self.prod_supplier_filter.addItem("كل الموردين", None)
        self.prod_supplier_filter.setFixedWidth(200)
        self.prod_supplier_filter.currentIndexChanged.connect(self.load_products)
        filters_layout.addWidget(self.prod_supplier_filter)
        
        # Date Filter
        self.prod_date_check = QCheckBox("تاريخ الفاتورة:")
        self.prod_date_check.toggled.connect(lambda: self.on_filter_changed())
        filters_layout.addWidget(self.prod_date_check)
        
        self.prod_date_filter = QDateEdit()
        self.prod_date_filter.setCalendarPopup(True)
        self.prod_date_filter.setDate(datetime.now().date())
        self.prod_date_filter.setEnabled(False)
        self.prod_date_filter.dateChanged.connect(self.load_products)
        self.prod_date_check.toggled.connect(self.prod_date_filter.setEnabled)
        filters_layout.addWidget(self.prod_date_filter)
        
        # Reset Filter Button
        self.reset_filters_btn = QPushButton("🔄 إعادة ضبط")
        self.reset_filters_btn.setStyleSheet(get_button_style('outline'))
        self.reset_filters_btn.clicked.connect(self.reset_advanced_filters)
        filters_layout.addWidget(self.reset_filters_btn)
        
        filters_layout.addStretch()
        layout.addLayout(filters_layout)
        
        # 4. Products Table (Sortable + Alternating Colors)
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(9) # 7 Visible + 2 Hidden (Supplier, Date)
        self.products_table.setHorizontalHeaderLabels(
            ["الرمز", "الاسم", "الفئة", "سعر الشراء", "سعر البيع", "الوحدة", "الإجراءات", "", ""]
        )
        # تخصيص عرض الأعمدة لجعل عمود الاسم أكبر
        self.products_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # الرمز
        self.products_table.setColumnWidth(0, 150)  # الرمز (أعرض)
        self.products_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # الاسم (يأخذ المساحة المتبقية)
        self.products_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # الفئة
        self.products_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # سعر الشراء
        self.products_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # سعر البيع
        self.products_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # الوحدة
        self.products_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.products_table.setColumnWidth(6, 250)  # الإجراءات (أعرض للأزرار العريضة)
        self.products_table.setColumnHidden(7, True) # Hidden Supplier
        self.products_table.setColumnHidden(8, True) # Hidden Date
        self.products_table.verticalHeader().setDefaultSectionSize(60)  # ارتفاع الصف (طبيعي)
        self.products_table.verticalHeader().setVisible(False)
        self.products_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.products_table.setStyleSheet(TABLE_STYLE)
        
        # Enable sorting
        self.products_table.setSortingEnabled(True)
        

        # Enable alternating row colors
        self.products_table.setAlternatingRowColors(True)
        
        # إخفاء عمود سعر الشراء إذا لم يكن مديراً للنظام
        if not getattr(self, 'can_view_buy_price', False):
            self.products_table.setColumnHidden(3, True)
            
        layout.addWidget(self.products_table, 1)
        
        # 5. Pagination Controls
        pagination_layout = QHBoxLayout()
        pagination_layout.setSpacing(10)
        
        self.products_results_label = QLabel("عدد النتائج: 0 منتج")
        self.products_results_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-weight: bold; font-size: 13px;")
        pagination_layout.addWidget(self.products_results_label)
        
        pagination_layout.addStretch()
        
        self.prev_page_btn = QPushButton("◀ السابق")
        self.prev_page_btn.setStyleSheet(get_button_style('outline') + "min-height: 35px; padding: 5px 20px;")
        self.prev_page_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.prev_page_btn.clicked.connect(self.previous_products_page)
        pagination_layout.addWidget(self.prev_page_btn)
        
        self.products_page_label = QLabel("صفحة 1 من 1")
        self.products_page_label.setStyleSheet(f"color: {COLORS['text_primary']}; font-weight: bold; padding: 0 15px;")
        pagination_layout.addWidget(self.products_page_label)
        
        self.next_page_btn = QPushButton("التالي ▶")
        self.next_page_btn.setStyleSheet(get_button_style('outline') + "min-height: 35px; padding: 5px 20px;")
        self.next_page_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.next_page_btn.clicked.connect(self.next_products_page)
        pagination_layout.addWidget(self.next_page_btn)
        
        layout.addLayout(pagination_layout)
        
        # Pagination state
        self.current_products_page = 1
        self.products_per_page = 50
        self.all_products = []
        
        scroll.setWidget(container)
        return scroll
    
    # ... (Inventory tab methods skipped, keeping existing structure) ...


    
    def create_inventory_tab(self):
        """إنشاء تبويب المخزون"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 1. Header Card with Inventory Count
        from PyQt6.QtWidgets import QFrame
        header_card = QFrame()
        header_card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['bg_main']}, stop:1 {COLORS['primary']});
                border: 2px solid {COLORS['border']};
                border-radius: 16px;
                padding: 20px;
            }}
        """)
        header_layout = QVBoxLayout(header_card)
        
        title = QLabel("📊 إدارة المخزون")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLORS['text_light']}; border: none; background: transparent;")
        header_layout.addWidget(title)
        
        self.inventory_count_label = QLabel("إجمالي الأصناف: 0 صنف")
        self.inventory_count_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 14px; border: none; background: transparent;")
        header_layout.addWidget(self.inventory_count_label)
        
        layout.addWidget(header_card)
        
        # 2. Store Filter Row
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        store_label = QLabel("🏪 الفرع:")
        store_label.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        store_label.setStyleSheet(f"color: {COLORS['text_primary']};")
        filter_layout.addWidget(store_label)
        
        self.store_filter_combo = QComboBox()
        self.store_filter_combo.setStyleSheet(INPUT_STYLE + "min-width: 250px;")
        self.store_filter_combo.setFont(QFont("Arial", 12))
        self.load_stores_to_combo()
        self.store_filter_combo.currentIndexChanged.connect(self.load_inventory)
        filter_layout.addWidget(self.store_filter_combo)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # 3. Search Bar (Separate Row)
        self.inventory_search_input = QLineEdit()
        self.inventory_search_input.setPlaceholderText("🔍 ابحث عن صنف بالاسم أو الكود...")
        self.inventory_search_input.setFont(QFont("Arial", 13))
        self.inventory_search_input.setStyleSheet(INPUT_STYLE + "min-height: 50px;")
        self.inventory_search_input.textChanged.connect(self.filter_inventory)
        layout.addWidget(self.inventory_search_input)
        
        # --- Advanced Filters Row ---
        inv_filters_layout = QHBoxLayout()
        inv_filters_layout.setSpacing(10)
        
        inv_filters_layout.addWidget(QLabel("المورد:"))
        self.inv_supplier_filter = QComboBox()
        self.inv_supplier_filter.addItem("كل الموردين", None)
        self.inv_supplier_filter.setFixedWidth(200)
        self.inv_supplier_filter.currentIndexChanged.connect(self.load_inventory)
        inv_filters_layout.addWidget(self.inv_supplier_filter)
        
        self.inv_date_check = QCheckBox("تاريخ الفاتورة:")
        self.inv_date_check.toggled.connect(lambda: self.load_inventory())
        inv_filters_layout.addWidget(self.inv_date_check)
        
        self.inv_date_filter = QDateEdit()
        self.inv_date_filter.setCalendarPopup(True)
        self.inv_date_filter.setDate(datetime.now().date())
        self.inv_date_filter.setEnabled(False)
        self.inv_date_filter.dateChanged.connect(self.load_inventory)
        self.inv_date_check.toggled.connect(self.inv_date_filter.setEnabled)
        inv_filters_layout.addWidget(self.inv_date_filter)
        
        inv_filters_layout.addStretch()
        layout.addLayout(inv_filters_layout)
        
        # 4. Inventory Table (Sortable + Alternating Colors)
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(5)
        self.inventory_table.setHorizontalHeaderLabels(
            ["رمز المنتج", "اسم المنتج", "الكمية الحالية", "الحد الأدنى", "الإجراءات"]
        )
        # تخصيص عرض الأعمدة لجعل عمود الاسم أكبر
        self.inventory_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # رمز المنتج
        self.inventory_table.setColumnWidth(0, 150)  # رمز المنتج (أعرض)
        self.inventory_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # اسم المنتج (يأخذ المساحة المتبقية)
        self.inventory_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # الكمية الحالية
        self.inventory_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # الحد الأدنى
        self.inventory_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.inventory_table.setColumnWidth(4, 300)  # الإجراءات (أعرض للأزرار العريضة)
        self.inventory_table.verticalHeader().setDefaultSectionSize(60)  # ارتفاع الصف (طبيعي)
        self.inventory_table.verticalHeader().setVisible(False)
        self.inventory_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.inventory_table.setStyleSheet(TABLE_STYLE)
        
        # Enable sorting
        self.inventory_table.setSortingEnabled(True)
        
        # Enable alternating row colors
        self.inventory_table.setAlternatingRowColors(True)
        
        layout.addWidget(self.inventory_table, 1)
        
        # 5. Pagination Controls
        pagination_layout = QHBoxLayout()
        pagination_layout.setSpacing(10)
        
        self.inventory_results_label = QLabel("عدد النتائج: 0 صنف")
        self.inventory_results_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-weight: bold; font-size: 13px;")
        pagination_layout.addWidget(self.inventory_results_label)
        
        pagination_layout.addStretch()
        
        self.prev_inventory_page_btn = QPushButton("◀ السابق")
        self.prev_inventory_page_btn.setStyleSheet(get_button_style('outline') + "min-height: 35px; padding: 5px 20px;")
        self.prev_inventory_page_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.prev_inventory_page_btn.clicked.connect(self.previous_inventory_page)
        pagination_layout.addWidget(self.prev_inventory_page_btn)
        
        self.inventory_page_label = QLabel("صفحة 1 من 1")
        self.inventory_page_label.setStyleSheet(f"color: {COLORS['text_primary']}; font-weight: bold; padding: 0 15px;")
        pagination_layout.addWidget(self.inventory_page_label)
        
        self.next_inventory_page_btn = QPushButton("التالي ▶")
        self.next_inventory_page_btn.setStyleSheet(get_button_style('outline') + "min-height: 35px; padding: 5px 20px;")
        self.next_inventory_page_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.next_inventory_page_btn.clicked.connect(self.next_inventory_page)
        pagination_layout.addWidget(self.next_inventory_page_btn)
        
        layout.addLayout(pagination_layout)
        
        # Pagination state
        self.current_inventory_page = 1
        self.inventory_per_page = 50
        self.all_inventory = []
        
        scroll.setWidget(container)
        return scroll

    def create_barcode_tab(self):
        """إنشاء تبويب طباعة الباركود بتصميم احترافي منقسم"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        container = QWidget()
        main_layout = QHBoxLayout(container)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(20)
        
        # --- الجانب الأيمن: التحكم والبحث (60% من العرض) ---
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        # 1. قسم صنف جديد
        new_item_card = QWidget()
        new_item_card.setObjectName("ProductCard")
        new_item_card.setStyleSheet(CARD_STYLE)
        new_item_layout = QVBoxLayout(new_item_card)
        
        new_item_title = QLabel("📦 صنف جديد / مخصص")
        new_item_title.setStyleSheet("color: " + COLORS['accent'] + "; font-size: 18px; font-weight: bold;")
        new_item_layout.addWidget(new_item_title)
        
        form_layout = QFormLayout()
        self.new_barcode_code = QLineEdit()
        self.new_barcode_code.setPlaceholderText("أدخل كود الصنف...")
        self.new_barcode_code.textChanged.connect(self.update_live_preview)
        
        self.new_barcode_name = QLineEdit()
        self.new_barcode_name.setPlaceholderText("أدخل اسم الصنف...")
        self.new_barcode_name.textChanged.connect(self.update_live_preview)

        self.new_barcode_price = QLineEdit()
        self.new_barcode_price.setPlaceholderText("أدخل السعر...")
        self.new_barcode_price.textChanged.connect(self.update_live_preview)
        
        form_layout.addRow("الكود:", self.new_barcode_code)
        form_layout.addRow("الاسم:", self.new_barcode_name)
        form_layout.addRow("السعر:", self.new_barcode_price)
        new_item_layout.addLayout(form_layout)
        right_layout.addWidget(new_item_card)
        
        # 2. قسم البحث والأصناف الحالية
        existing_group = QWidget()
        existing_layout = QVBoxLayout(existing_group)
        
        search_label = QLabel("🔍 البحث في الأصناف الحالية")
        search_label.setStyleSheet(LABEL_STYLE_TITLE)
        existing_layout.addWidget(search_label)
        
        self.barcode_search_input = QLineEdit()
        self.barcode_search_input.setPlaceholderText("بحث بالحروف أو الكود...")
        self.barcode_search_input.textChanged.connect(self.filter_barcode_products)
        existing_layout.addWidget(self.barcode_search_input)
        
        # --- Advanced Filters Row ---
        barcode_filters_layout = QHBoxLayout()
        barcode_filters_layout.setSpacing(10)
        
        barcode_filters_layout.addWidget(QLabel("المورد:"))
        self.barcode_supplier_filter = QComboBox()
        self.barcode_supplier_filter.addItem("كل الموردين", None)
        self.barcode_supplier_filter.setFixedWidth(200)
        self.barcode_supplier_filter.currentIndexChanged.connect(self.load_barcode_products)
        barcode_filters_layout.addWidget(self.barcode_supplier_filter)
        
        self.barcode_date_check = QCheckBox("تاريخ الفاتورة:")
        self.barcode_date_check.toggled.connect(lambda: self.load_barcode_products())
        barcode_filters_layout.addWidget(self.barcode_date_check)
        
        self.barcode_date_filter = QDateEdit()
        self.barcode_date_filter.setCalendarPopup(True)
        self.barcode_date_filter.setDate(datetime.now().date())
        self.barcode_date_filter.setEnabled(False)
        self.barcode_date_filter.dateChanged.connect(self.load_barcode_products)
        self.barcode_date_check.toggled.connect(self.barcode_date_filter.setEnabled)
        barcode_filters_layout.addWidget(self.barcode_date_filter)
        
        existing_layout.addLayout(barcode_filters_layout)
        
        self.barcode_table = QTableWidget()
        self.barcode_table.setColumnCount(4) # Added hidden price column
        self.barcode_table.verticalHeader().setDefaultSectionSize(45)
        self.barcode_table.verticalHeader().setVisible(False)
        self.barcode_table.setHorizontalHeaderLabels(["الرمز", "الاسم", "إجراء", ""])
        self.barcode_table.setColumnHidden(3, True) # Hide price
        self.barcode_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.barcode_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.barcode_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.barcode_table.itemSelectionChanged.connect(self.on_barcode_selection_changed)
        existing_layout.addWidget(self.barcode_table)
        
        right_layout.addWidget(existing_group, 1)
        
        # --- الجانب الأيسر: المعاينة المباشرة والطباعة (40% من العرض) ---
        left_panel = QWidget()
        left_panel.setFixedWidth(350)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(15)
        
        preview_card = QWidget()
        preview_card.setObjectName("ProductCard")
        preview_card.setStyleSheet(CARD_STYLE)
        preview_vbox = QVBoxLayout(preview_card)
        
        preview_title = QLabel("👀 المعاينة المباشرة")
        preview_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        preview_title.setStyleSheet(LABEL_STYLE_TITLE)
        preview_vbox.addWidget(preview_title)
        
        # حاوية صورة الباركود
        self.barcode_preview_label = QLabel("أدخل بيانات للمعاينة")
        self.barcode_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.barcode_preview_label.setMinimumHeight(150)
        self.barcode_preview_label.setStyleSheet(f"background-color: white; border: 1px solid {COLORS['border']}; border-radius: 8px; color: #333;")
        preview_vbox.addWidget(self.barcode_preview_label)
        
        # تفاصيل الصنف المعروض
        self.preview_info_label = QLabel("")
        self.preview_info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_info_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        preview_vbox.addWidget(self.preview_info_label)
        
        # مقاس الملصق وعدد النسخ واختيار الطابعة
        settings_layout = QVBoxLayout()
        
        # السطر الأول: المقاس والنسخ
        row1 = QHBoxLayout()
        size_layout = QVBoxLayout()
        size_layout.addWidget(QLabel("📏 مقاس الملصق:"))
        self.label_size_combo = QComboBox()
        self.label_size_combo.addItems(["50x30", "40x25"])  # 50x30 is now default
        self.label_size_combo.currentTextChanged.connect(self.update_live_preview)
        size_layout.addWidget(self.label_size_combo)
        row1.addLayout(size_layout)
        
        copies_layout = QVBoxLayout()
        copies_layout.addWidget(QLabel("🔢 عدد النسخ:"))
        self.barcode_copies_spin = QSpinBox()
        self.barcode_copies_spin.setRange(1, 100)
        self.barcode_copies_spin.setValue(1)
        copies_layout.addWidget(self.barcode_copies_spin)
        row1.addLayout(copies_layout)
        settings_layout.addLayout(row1)
        
        # السطر الثاني: اختيار الطابعة
        printer_layout = QVBoxLayout()
        printer_layout.addWidget(QLabel("🖨️ اختر الطابعة:"))
        self.printer_selector = QComboBox()
        # جلب قائمة الطابعات
        printers = [p.printerName() for p in QPrinterInfo.availablePrinters()]
        self.printer_selector.addItems(printers)
        # تحديد الطابعة الافتراضية
        default_printer = QPrinterInfo.defaultPrinter().printerName()
        index = self.printer_selector.findText(default_printer)
        if index >= 0:
            self.printer_selector.setCurrentIndex(index)
        printer_layout.addWidget(self.printer_selector)
        settings_layout.addLayout(printer_layout)
        
        preview_vbox.addLayout(settings_layout)
        
        # زر الطباعة الكبير
        self.print_main_btn = QPushButton("🖨️ طباعة الآن")
        self.print_main_btn.setStyleSheet(get_button_style(COLORS['success']))
        self.print_main_btn.clicked.connect(self.print_highlighted_barcode)
        preview_vbox.addWidget(self.print_main_btn)
        
        preview_vbox.addStretch()
        left_layout.addWidget(preview_card)
        left_layout.addStretch()
        
        main_layout.addWidget(right_panel, 6)
        main_layout.addWidget(left_panel, 4)
        
        scroll.setWidget(container)
        return scroll

    def update_live_preview(self):
        """تحديث المعاينة المباشرة بناءً على الحقول اليدوية والمقاس المختار"""
        code = self.new_barcode_code.text().strip()
        name = self.new_barcode_name.text().strip()
        price = self.new_barcode_price.text().strip()
        size = self.label_size_combo.currentText()
        
        if code:
            self.generate_and_show_preview(code, name, size, price)
            # مسح الاختيار من الجدول إذا تم البدء في الكتابة يدوياً
            self.barcode_table.clearSelection()
        else:
            self.barcode_preview_label.setText("أدخل بيانات للمعاينة")
            self.preview_info_label.setText("")

    def on_barcode_selection_changed(self):
        """تحديث المعاينة عند اختيار صنف من الجدول"""
        selected_items = self.barcode_table.selectedItems()
        if selected_items:
            # الصف الأول هو الكود، الثاني هو الاسم
            row = selected_items[0].row()
            code = self.barcode_table.item(row, 0).text()
            name = self.barcode_table.item(row, 1).text()
            
            # حفظ السعر للطباعة
            price_item = self.barcode_table.item(row, 3)
            price = price_item.text() if price_item else "0.00"
            self.current_barcode_price = price
            
            # تحديث الحقول اليدوية لتعكس المختار (اختياري، يسهل التعديل)
            self.new_barcode_code.blockSignals(True)
            self.new_barcode_name.blockSignals(True)
            self.new_barcode_price.blockSignals(True)
            self.new_barcode_code.setText(code)
            self.new_barcode_name.setText(name)
            self.new_barcode_price.setText(price)
            self.new_barcode_code.blockSignals(False)
            self.new_barcode_name.blockSignals(False)
            self.new_barcode_price.blockSignals(False)
            
            size = self.label_size_combo.currentText()
            self.generate_and_show_preview(code, name, size, price)

    def generate_and_show_preview(self, code, name, size="40x25", price=None):
        """توليد صورة باركود للمعاينة بمقاس محدد"""
        try:
            from utils.barcode_service import BarcodeService
            from PyQt6.QtGui import QPixmap
            
            # توليد الباركود في مجلد مؤقت بالمقاس المختار
            path = BarcodeService.generate_barcode(code, name, label_size=size)
            if path and os.path.exists(path):
                pixmap = QPixmap(path)
                # تصغير الصورة لتناسب المساحة مع الحفاظ على التناسب
                self.barcode_preview_label.setPixmap(pixmap.scaled(
                    300, 150, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation
                ))
                info_text = f"{name}\n({code})"
                if price:
                    info_text += f"\nالسعر: {float(price):.2f} ج.م"
                self.preview_info_label.setText(info_text)
        except Exception as e:
            print(f"Preview error: {e}")

    def print_highlighted_barcode(self):
        """طباعة الصنف المعروض حالياً في المعاينة بالعدد المطلوب"""
        code = self.new_barcode_code.text().strip()
        name = self.new_barcode_name.text().strip()
        num_copies = self.barcode_copies_spin.value()
        size = self.label_size_combo.currentText()
        selected_printer = self.printer_selector.currentText()
        
        if not code:
            QMessageBox.warning(self, "تنبيه", "يرجى اختيار صنف أو إدخال بيانات أولاً")
            return
            
        if not selected_printer:
            QMessageBox.warning(self, "تنبيه", "يرجى اختيار طابعة أولاً")
            return

        # تنفيذ الطباعة المباشرة بعدد المرات المطلوب وبالمقاس المختار
        success_count = 0
        
        # استخدام السعر المحفوظ إذا كان متاحاً، وإلا فارغ
        price = self.new_barcode_price.text().strip()
        store = self.user_info.get('store_name', '')
        
        for _ in range(num_copies):
            if PrinterService.print_barcode_direct(name, code, selected_printer, label_size=size, price=price, store_name=store):
                success_count += 1
        
        if success_count > 0:
            QMessageBox.information(self, "نجاح", f"تم إرسال {success_count} نسخة للطباعة")
        else:
            QMessageBox.critical(self, "خطأ", "فشلت عملية الطباعة")

    def filter_barcode_products(self):
        """تصفية قائمة المنتجات في تبويب الباركود"""
        self.load_barcode_products()

    def load_barcode_products(self):
        """تحميل المنتجات في جدول الباركود مع دعم الفلاتر"""
        try:
            # Check for Advanced Filters
            supplier_id = self.barcode_supplier_filter.currentData()
            invoice_date = None
            search_text = self.barcode_search_input.text().strip() if hasattr(self, 'barcode_search_input') else ""
            if self.barcode_date_check.isChecked():
                invoice_date = self.barcode_date_filter.date().toPyDate().strftime('%Y-%m-%d')
            
            if supplier_id or invoice_date:
                # Use current store or 1
                store_id = self.user_info.get('store_id') or 1
                products = self.db.get_advanced_product_search(store_id, supplier_id, invoice_date)
                if search_text:
                    search_l = search_text.lower()
                    products = [
                        p for p in products
                        if search_l in str(p.get('product_code', '')).lower()
                        or search_l in str(p.get('product_name', '')).lower()
                    ]
                products = products[:self._max_barcode_rows]
            else:
                cursor = self.db.cursor
                params = []
                query = """
                    SELECT id, product_code, product_name, sell_price
                    FROM products
                    WHERE is_active = TRUE
                """
                if search_text:
                    query += " AND (product_code LIKE %s OR product_name LIKE %s)"
                    like = f"%{search_text}%"
                    params.extend([like, like])
                query += " ORDER BY product_name LIMIT %s"
                params.append(self._max_barcode_rows)
                cursor.execute(query, tuple(params))
                products = cursor.fetchall()
            
            self.barcode_table.setRowCount(len(products))
            for row, product in enumerate(products):
                self.barcode_table.setItem(row, 0, QTableWidgetItem(product['product_code']))
                self.barcode_table.setItem(row, 1, QTableWidgetItem(product['product_name']))
                
                # إضافة السعر في عمود مخفي
                price_str = f"{product['sell_price']:.2f}"
                self.barcode_table.setItem(row, 3, QTableWidgetItem(price_str))
                
                # إنشاء حاوية لتوسيط الزر وتصغير حجمه
                btn_container = QWidget()
                btn_layout = QHBoxLayout(btn_container)
                btn_layout.setContentsMargins(0, 0, 0, 0)
                btn_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                
                print_btn = QPushButton("👁️ معاينة")
                # استخدام نمط مخصص للجدول (أصغر)
                print_btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {COLORS['info']};
                        color: white;
                        border-radius: 6px;
                        padding: 2px 15px;
                        font-weight: 500;
                        font-size: 13px;
                        min-height: 28px;
                        max-width: 100px;
                    }}
                    QPushButton:hover {{
                        background-color: #2563EB;
                    }}
                """)
                print_btn.clicked.connect(lambda checked, r=row: self.on_preview_btn_clicked(r))
                btn_layout.addWidget(print_btn)
                
                self.barcode_table.setCellWidget(row, 2, btn_container)
            
            # Dynamic Table Height
            row_height = 45
            header_height = self.barcode_table.horizontalHeader().height()
            content_height = header_height + (len(products) * row_height) + 10
            self.barcode_table.setFixedHeight(min(content_height, 1200))
            self._barcode_loaded = True
        except Exception as e:
            print(f"Error loading barcode products: {e}")

    def on_preview_btn_clicked(self, row):
        """التعامل مع النقر على زر المعاينة داخل الجدول"""
        self.barcode_table.selectRow(row)
        # استخراج البيانات مباشرة من الجدول
        code_item = self.barcode_table.item(row, 0)
        name_item = self.barcode_table.item(row, 1)
        if code_item and name_item:
            code = code_item.text()
            name = name_item.text()
            
            # تحديث الحقول اليدوية
            self.new_barcode_code.blockSignals(True)
            self.new_barcode_name.blockSignals(True)
            self.new_barcode_code.setText(code)
            self.new_barcode_name.setText(name)
            self.new_barcode_code.blockSignals(False)
            self.new_barcode_name.blockSignals(False)
            
            # عرض المعاينة
            size = self.label_size_combo.currentText()
            self.generate_and_show_preview(code, name, size)
    
    def populate_suppliers_combos(self):
        """تعبئة قوائم الموردين في جميع التبويبات"""
        try:
            cursor = self.db.cursor
            cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
            suppliers = cursor.fetchall()
            
            # Clear and add "All" option to each
            for combo in [self.prod_supplier_filter, self.inv_supplier_filter, self.barcode_supplier_filter]:
                combo.blockSignals(True)
                combo.clear()
                combo.addItem("كل الموردين", None)
                for s in suppliers:
                    combo.addItem(s['name'], s['id'])
                combo.blockSignals(False)
        except Exception as e:
            print(f"Error populating suppliers: {e}")

    def on_filter_changed(self):
        """التعامل مع تغيير حالة الفلتر"""
        # This is a generic handler that can be expanded if needed
        self.load_products()
        self.load_inventory()
        self.load_barcode_products()

    def reset_advanced_filters(self):
        """إعادة تعيين كافة الفلاتر لمسح البحث المتقدم"""
        # Products Tab
        self.prod_supplier_filter.blockSignals(True)
        self.prod_supplier_filter.setCurrentIndex(0)
        self.prod_supplier_filter.blockSignals(False)
        
        self.prod_date_check.blockSignals(True)
        self.prod_date_check.setChecked(False)
        self.prod_date_check.blockSignals(False)
        self.prod_date_filter.setEnabled(False)
        
        # Inventory Tab
        self.inv_supplier_filter.blockSignals(True)
        self.inv_supplier_filter.setCurrentIndex(0)
        self.inv_supplier_filter.blockSignals(False)
        
        self.inv_date_check.blockSignals(True)
        self.inv_date_check.setChecked(False)
        self.inv_date_check.blockSignals(False)
        self.inv_date_filter.setEnabled(False)
        
        # Barcode Tab
        self.barcode_supplier_filter.blockSignals(True)
        self.barcode_supplier_filter.setCurrentIndex(0)
        self.barcode_supplier_filter.blockSignals(False)
        
        self.barcode_date_check.blockSignals(True)
        self.barcode_date_check.setChecked(False)
        self.barcode_date_check.blockSignals(False)
        self.barcode_date_filter.setEnabled(False)
        
        self.search_input.clear()
        self.load_products()
        self.load_inventory()
        self.load_barcode_products()

    def load_products(self):
        """تحميل المنتجات مع دعم الفلاتر المتقدمة والترقيم"""
        try:
            # Temporarily disable sorting while loading
            self.products_table.setSortingEnabled(False)
            
            # Check for Advanced Filters + Search
            supplier_id = self.prod_supplier_filter.currentData()
            invoice_date = None
            search_text = self.search_input.text().strip() if hasattr(self, 'search_input') else ""
            if self.prod_date_check.isChecked():
                invoice_date = self.prod_date_filter.date().toPyDate().strftime('%Y-%m-%d')
            
            # 1. Fetch Data
            if supplier_id or invoice_date:
                # Advanced Search (in-memory pagination after filtered query)
                store_id = self.user_info.get('store_id') or 1
                self.all_products = self.db.get_advanced_product_search(store_id, supplier_id, invoice_date)
                if search_text:
                    s = search_text.lower()
                    self.all_products = [
                        p for p in self.all_products
                        if s in str(p.get('product_code', '')).lower()
                        or s in str(p.get('product_name', '')).lower()
                        or s in str(p.get('supplier_name', '')).lower()
                    ]
                total_products = len(self.all_products)
                self._products_total_count = total_products
                total_pages = max(1, (total_products + self.products_per_page - 1) // self.products_per_page)
                if self.current_products_page > total_pages:
                    self.current_products_page = total_pages
                start_idx = (self.current_products_page - 1) * self.products_per_page
                end_idx = min(start_idx + self.products_per_page, total_products)
                page_products = self.all_products[start_idx:end_idx]
            else:
                # Normal View - Server-side pagination (fast for large datasets)
                cursor = self.db.cursor
                where_clause = " WHERE p.is_active = TRUE "
                params = []
                if search_text:
                    where_clause += " AND (p.product_code LIKE %s OR p.product_name LIKE %s OR COALESCE(s.name, '') LIKE %s) "
                    like = f"%{search_text}%"
                    params.extend([like, like, like])

                count_query = f"""
                    SELECT COUNT(*) AS total_count
                    FROM products p
                    LEFT JOIN suppliers s ON p.supplier_id = s.id
                    {where_clause}
                """
                cursor.execute(count_query, tuple(params))
                total_products = int((cursor.fetchone() or {}).get('total_count', 0))
                self._products_total_count = total_products

                total_pages = max(1, (total_products + self.products_per_page - 1) // self.products_per_page)
                if self.current_products_page > total_pages:
                    self.current_products_page = total_pages

                offset = (self.current_products_page - 1) * self.products_per_page
                data_query = f"""
                    SELECT p.id, p.product_code, p.product_name, c.category_name,
                           p.buy_price, p.sell_price, p.unit, s.name as supplier_name
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.id
                    LEFT JOIN suppliers s ON p.supplier_id = s.id
                    {where_clause}
                    ORDER BY p.product_name
                    LIMIT %s OFFSET %s
                """
                page_params = params + [self.products_per_page, offset]
                cursor.execute(data_query, tuple(page_params))
                page_products = cursor.fetchall()
                self.all_products = page_products
            
            # 2. Unified Table Configuration (8 Columns)
            self.products_table.setColumnCount(8)
            self.products_table.setHorizontalHeaderLabels([
                "الرمز", "الاسم", "الفئة", "سعر الشراء", "سعر البيع", "الوحدة", "المورد", "الإجراءات"
            ])
            for i in range(8): 
                self.products_table.setColumnHidden(i, False)
            
            # Standard Widths
            self.products_table.setColumnWidth(1, 250) # الاسم
            self.products_table.setColumnWidth(7, 150) # الإجراءات
            self.products_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
            
            actions_col = 7
            
            total_products = self._products_total_count if self._products_total_count else len(page_products)
            
            # Update product count label
            if hasattr(self, 'products_count_label'):
                self.products_count_label.setText(f"إجمالي المنتجات: {total_products} منتج")
            
            # Calculate pagination (already computed in fetch block, keep as fallback)
            total_pages = max(1, (total_products + self.products_per_page - 1) // self.products_per_page)
            
            # Update table
            self.products_table.clearContents()
            self.products_table.setRowCount(len(page_products))
            
            # Check permissions for edit/delete buttons
            role_name = self.user_info.get('role_name', '').lower()
            role_id = self.user_info.get('role_id')
            allowed_roles = ['admin', 'system admin', 'مدير عام', 'developer']
            can_edit = any(role in role_name for role in allowed_roles) or (role_id in [99, 1])
            
            for row, product in enumerate(page_products):
                # Unified Mapping: Code, Name, Cat, Buy, Sell, Unit, Supplier, Actions
                self.products_table.setItem(row, 0, QTableWidgetItem(str(product.get('product_code', ''))))
                self.products_table.setItem(row, 1, QTableWidgetItem(str(product.get('product_name', ''))))
                self.products_table.setItem(row, 2, QTableWidgetItem(product.get('category_name', "") or "غير مصنف"))
                
                buy_price = float(product.get('buy_price', 0))
                sell_price = float(product.get('sell_price', 0))
                
                buy_item = QTableWidgetItem(f"{buy_price:,.2f}")
                if not self.can_view_buy_price:
                    buy_item = QTableWidgetItem("****")
                    
                self.products_table.setItem(row, 3, buy_item)
                self.products_table.setItem(row, 4, QTableWidgetItem(f"{sell_price:,.2f}"))
                self.products_table.setItem(row, 5, QTableWidgetItem(product.get('unit', "قطعة") or "قطعة"))
                self.products_table.setItem(row, 6, QTableWidgetItem(product.get('supplier_name', "غير محدد") or "غير محدد"))
                
                actions_col = 7
                
                # Actions column
                if can_edit:
                    self.products_table.setColumnHidden(actions_col, False)
                    actions_widget = QWidget()
                    actions_layout = QHBoxLayout(actions_widget)
                    actions_layout.setContentsMargins(2, 2, 2, 2)
                    actions_layout.setSpacing(3)
                    
                    edit_btn = QPushButton("✏️ تعديل")
                    edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                    edit_btn.setToolTip("تعديل بيانات المنتج")
                    edit_btn.setStyleSheet(f"""
                        QPushButton {{
                            background-color: {COLORS['info']};
                            color: white;
                            border-radius: 4px;
                            padding: 2px 8px;
                            font-weight: bold;
                            font-size: 12px;
                            min-height: 26px;
                            max-height: 30px;
                        }}
                        QPushButton:hover {{
                            background-color: #2563EB;
                        }}
                    """)
                    edit_btn.clicked.connect(lambda checked, p_id=product['id']: self.edit_product(p_id))
                    actions_layout.addWidget(edit_btn)
                    
                    delete_btn = QPushButton("🗑️")
                    delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                    delete_btn.setToolTip("حذف المنتج")
                    delete_btn.setStyleSheet(f"""
                        QPushButton {{
                            background-color: {COLORS['danger']};
                            color: white;
                            border-radius: 4px;
                            padding: 2px 6px;
                            font-size: 13px;
                            min-height: 26px;
                            max-height: 30px;
                            min-width: 35px;
                        }}
                        QPushButton:hover {{
                            background-color: #DC2626;
                        }}
                    """)
                    delete_btn.clicked.connect(lambda checked, p_id=product['id'], p_name=product['product_name']: self.delete_product(p_id, p_name))
                    actions_layout.addWidget(delete_btn)
                    
                    self.products_table.setCellWidget(row, actions_col, actions_widget)
                else:
                    item = QTableWidgetItem("🔒")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setToolTip("غير مصرح بالتعديل")
                    self.products_table.setItem(row, actions_col, item)
            
            # Update pagination controls
            if hasattr(self, 'products_page_label'):
                self.products_page_label.setText(f"صفحة {self.current_products_page} من {total_pages}")
            
            if hasattr(self, 'products_results_label'):
                self.products_results_label.setText(f"عدد النتائج: {len(page_products)} من {total_products} منتج")
            
            if hasattr(self, 'prev_page_btn'):
                self.prev_page_btn.setEnabled(self.current_products_page > 1)
            
            if hasattr(self, 'next_page_btn'):
                self.next_page_btn.setEnabled(self.current_products_page < total_pages)
            
            # Re-enable sorting
            self.products_table.setSortingEnabled(True)
            
            # Dynamic Table Height
            row_height = 60
            header_height = self.products_table.horizontalHeader().height()
            content_height = header_height + (len(page_products) * row_height) + 10
            self.products_table.setFixedHeight(min(content_height, 1200))
            self._products_loaded = True
        
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"خطأ في تحميل المنتجات: {str(e)}")

    
    def load_inventory(self):
        """تحميل المخزون مع دعم الفلاتر المتقدمة والترقيم"""
        try:
            # Temporarily disable sorting while loading
            self.inventory_table.setSortingEnabled(False)
            
            target_store_id = self.store_filter_combo.currentData()
            user_role_id = self.user_info.get('role_id', 0)
            user_store_id = self.user_info.get('store_id')
            
            # Check for Advanced Filters
            supplier_id = self.inv_supplier_filter.currentData()
            invoice_date = None
            search_text = self.inventory_search_input.text().strip() if hasattr(self, 'inventory_search_input') else ""
            if self.inv_date_check.isChecked():
                invoice_date = self.inv_date_filter.date().toPyDate().strftime('%Y-%m-%d')
            
            # تحديث ويدجت التحويلات الواردة إذا كان المستخدم أدمن
            if hasattr(self, 'incoming_transfers_widget'):
                self.incoming_transfers_widget.load_transfers(target_store_id)
            
            # معرفة ما إذا كان المتجر المحدد هو المخزن الرئيسي
            main_warehouse_id = self.db.get_main_warehouse_id()
            is_target_main = (target_store_id == main_warehouse_id)
            is_user_main_manager = (user_store_id == main_warehouse_id)
            
            # تحديد الصلاحيات
            can_edit_manual = False
            can_transfer = False
            
            # 1. Admin/Developer/Cashier (Full Control)
            if user_role_id in [1, 99, 3]:
                can_edit_manual = True
                can_transfer = True
            
            # 2. Manager
            elif user_role_id == 2:
                if is_target_main:
                    if is_user_main_manager:
                        can_transfer = True
                        can_edit_manual = False
                else:
                    if user_store_id == target_store_id:
                        can_transfer = True
            
            # Load Data
            if supplier_id or invoice_date:
                # Advanced Search (Filtered by Supplier/Date)
                self.all_inventory = self.db.get_advanced_product_search(target_store_id, supplier_id, invoice_date)
                if search_text:
                    s = search_text.lower()
                    self.all_inventory = [
                        it for it in self.all_inventory
                        if s in str(it.get('product_code', '')).lower()
                        or s in str(it.get('product_name', '')).lower()
                    ]
                total_inventory = len(self.all_inventory)
                self._inventory_total_count = total_inventory
                total_pages = max(1, (total_inventory + self.inventory_per_page - 1) // self.inventory_per_page)
                if self.current_inventory_page > total_pages:
                    self.current_inventory_page = total_pages
                start_idx = (self.current_inventory_page - 1) * self.inventory_per_page
                end_idx = min(start_idx + self.inventory_per_page, total_inventory)
                page_inventory = self.all_inventory[start_idx:end_idx]
            else:
                # Normal View (Server-side pagination for large datasets)
                cursor = self.db.cursor
                search_clause = ""
                search_params = []
                if search_text:
                    search_clause = " AND (p.product_code LIKE %s OR p.product_name LIKE %s)"
                    like = f"%{search_text}%"
                    search_params.extend([like, like])

                count_query = f"""
                    SELECT COUNT(*) AS total_count
                    FROM products p
                    WHERE p.is_active = TRUE
                    {search_clause}
                """
                cursor.execute(count_query, tuple(search_params))
                total_inventory = int((cursor.fetchone() or {}).get('total_count', 0))
                self._inventory_total_count = total_inventory

                total_pages = max(1, (total_inventory + self.inventory_per_page - 1) // self.inventory_per_page)
                if self.current_inventory_page > total_pages:
                    self.current_inventory_page = total_pages

                offset = (self.current_inventory_page - 1) * self.inventory_per_page
                data_query = f"""
                    SELECT p.id as product_id, p.product_code, p.product_name, 
                           COALESCE(i.quantity_in_stock, 0) as quantity_in_stock, 
                           COALESCE(i.minimum_quantity, 0) as minimum_quantity
                    FROM products p
                    LEFT JOIN product_inventory i ON p.id = i.product_id AND i.store_id = %s
                    WHERE p.is_active = TRUE
                    {search_clause}
                    ORDER BY p.product_name
                    LIMIT %s OFFSET %s
                """
                data_params = [target_store_id] + search_params + [self.inventory_per_page, offset]
                cursor.execute(data_query, tuple(data_params))
                page_inventory = cursor.fetchall()
                self.all_inventory = page_inventory
                
            total_inventory = self._inventory_total_count if self._inventory_total_count else len(page_inventory)
            
            # Update inventory count label
            if hasattr(self, 'inventory_count_label'):
                self.inventory_count_label.setText(f"إجمالي الأصناف: {total_inventory} صنف")
            
            # Calculate pagination (already computed in fetch block, keep as fallback)
            total_pages = max(1, (total_inventory + self.inventory_per_page - 1) // self.inventory_per_page)
            
            # Determine correct columns and count
            if supplier_id or invoice_date:
                # Advanced View: Code(0), Name(1), Current(2), Pur(3), Sold(4), Supp(5), Date(6), Actions(7)
                self.inventory_table.setColumnCount(8)
                self.inventory_table.setHorizontalHeaderLabels([
                    "الرمز", "الاسم", "المخزون الحالي", "كمية الشراء", "المباع منها", "المورد", "التاريخ", "الإجراءات"
                ])
                for i in range(8): self.inventory_table.setColumnHidden(i, False)
                
                # Adjust Column Widths
                self.inventory_table.setColumnWidth(3, 100) # Purchase Qty
                self.inventory_table.setColumnWidth(4, 90)  # Sold Qty (Reduced)
                self.inventory_table.setColumnWidth(7, 350) # Actions (Increased)
                self.inventory_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
                
                actions_col = 7
            else:
                # Normal View: Code(0), Name(1), Current(2), Min(3), Actions(4)
                self.inventory_table.setColumnCount(5)
                self.inventory_table.setHorizontalHeaderLabels([
                    "رمز المنتج", "اسم المنتج", "الكمية الحالية", "الحد الأدنى", "الإجراءات"
                ])
                for i in range(5): self.inventory_table.setColumnHidden(i, False)
                
                # Restore Normal Widths
                self.inventory_table.setColumnWidth(4, 300)
                self.inventory_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
                
                actions_col = 4

            # Update table
            self.inventory_table.clearContents()
            self.inventory_table.setRowCount(len(page_inventory))
            
            # Show/Hide Actions column based on permissions
            if can_edit_manual or can_transfer:
                self.inventory_table.setColumnHidden(actions_col, False)
            else:
                self.inventory_table.setColumnHidden(actions_col, True)
                
            for row, item in enumerate(page_inventory):
                # Using a dictionary to determine column mapping
                if supplier_id or invoice_date:
                    # Advanced Columns: Code, Name, Stock, PurQty, SoldQty, Supplier, Date, Actions
                    self.inventory_table.setItem(row, 0, QTableWidgetItem(item['product_code']))
                    self.inventory_table.setItem(row, 1, QTableWidgetItem(item['product_name']))
                    self.inventory_table.setItem(row, 2, QTableWidgetItem(str(item['quantity_in_stock'])))
                    self.inventory_table.setItem(row, 3, QTableWidgetItem(str(item.get('purchase_qty', 0))))
                    self.inventory_table.setItem(row, 4, QTableWidgetItem(str(item.get('sold_qty_since_purchase', 0))))
                    self.inventory_table.setItem(row, 5, QTableWidgetItem(item.get('supplier_name', "")))
                    self.inventory_table.setItem(row, 6, QTableWidgetItem(str(item.get('invoice_date', ""))))
                    actions_col = 7
                else:
                    # Normal Columns: Code, Name, Stock, Min, Actions
                    # تلوين الصفوف ذات المخزون المنخفض (Normal View Only)
                    is_low = item['quantity_in_stock'] <= item.get('minimum_quantity', 0)
                    
                    code_item = QTableWidgetItem(item['product_code'])
                    name_item = QTableWidgetItem(item['product_name'])
                    stock_item = QTableWidgetItem(str(item['quantity_in_stock']))
                    min_item = QTableWidgetItem(str(item.get('minimum_quantity', 0)))
                    
                    if is_low:
                        danger_color = QColor(COLORS['danger'])
                        code_item.setForeground(danger_color)
                        name_item.setForeground(danger_color)
                        stock_item.setForeground(danger_color)
                        min_item.setForeground(danger_color)
                        f = name_item.font()
                        f.setBold(True)
                        name_item.setFont(f)

                    self.inventory_table.setItem(row, 0, code_item)
                    self.inventory_table.setItem(row, 1, name_item)
                    self.inventory_table.setItem(row, 2, stock_item)
                    self.inventory_table.setItem(row, 3, min_item)
                    actions_col = 4
                
                # إنشاء حاوية الأزرار بتصميم محسّن
                if can_edit_manual or can_transfer:
                    widget = QWidget()
                    btn_layout = QHBoxLayout(widget)
                    btn_layout.setContentsMargins(2, 2, 2, 2)
                    btn_layout.setSpacing(3)
                    
                    # زر التعديل اليدوي (Only Admin)
                    if can_edit_manual:
                        edit_btn = QPushButton("🔧 تعديل")
                        edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                        edit_btn.setToolTip("تعديل الكمية يدوياً (جرد)")
                        edit_btn.setStyleSheet(f"""
                            QPushButton {{
                                background-color: {COLORS['danger']};
                                color: white;
                                border-radius: 4px;
                                padding: 2px 8px;
                                font-weight: bold;
                                font-size: 12px;
                                min-height: 26px;
                                max-height: 30px;
                            }}
                            QPushButton:hover {{
                                background-color: #DC2626;
                            }}
                        """)
                        edit_btn.clicked.connect(lambda ch, pid=item.get('product_id', item.get('id')): self.open_manual_edit(pid, target_store_id))
                        btn_layout.addWidget(edit_btn)
                    
                    # زر التحويل (Admin or Manager)
                    if can_transfer:
                        trans_btn = QPushButton("🚚 تحويل")
                        trans_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                        trans_btn.setToolTip("تحويل مخزون لفرع آخر")
                        trans_btn.setStyleSheet(f"""
                            QPushButton {{
                                background-color: {COLORS['success']};
                                color: white;
                                border-radius: 4px;
                                padding: 2px 8px;
                                font-weight: bold;
                                font-size: 12px;
                                min-height: 26px;
                                max-height: 30px;
                            }}
                            QPushButton:hover {{
                                background-color: {COLORS['accent_hover']};
                            }}
                        """)
                        trans_btn.clicked.connect(lambda ch, pid=item.get('product_id', item.get('id')), pname=item['product_name']: self.open_transfer(pid, target_store_id, pname))
                        btn_layout.addWidget(trans_btn)
                        
                    self.inventory_table.setCellWidget(row, actions_col, widget)
            
            # Update pagination controls
            if hasattr(self, 'inventory_page_label'):
                self.inventory_page_label.setText(f"صفحة {self.current_inventory_page} من {total_pages}")
            
            if hasattr(self, 'inventory_results_label'):
                self.inventory_results_label.setText(f"عدد النتائج: {len(page_inventory)} من {total_inventory} صنف")
            
            if hasattr(self, 'prev_inventory_page_btn'):
                self.prev_inventory_page_btn.setEnabled(self.current_inventory_page > 1)
            
            if hasattr(self, 'next_inventory_page_btn'):
                self.next_inventory_page_btn.setEnabled(self.current_inventory_page < total_pages)
            
            # Re-enable sorting
            self.inventory_table.setSortingEnabled(True)
            
            # Dynamic Table Height
            row_height = 60
            header_height = self.inventory_table.horizontalHeader().height()
            content_height = header_height + (len(page_inventory) * row_height) + 10
            self.inventory_table.setFixedHeight(min(content_height, 1200))
            self._inventory_loaded = True

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"خطأ في تحميل المخزون: {str(e)}")


    
    def load_stores_to_combo(self):
        """تحميل الفروع في القائمة المنسدلة"""
        try:
            cursor = self.db.cursor
            # المستخدم العادي (Branch Manager) يشوف بس متجره؟ والا يشوف الكل (View Only)؟
            # الطلب ما حدد، بس قال "لا استطيع تعديل كميات المنتجات فقط عرض" لفرع الاسكندرية
            # إذن يمكنه عرض الكميات.
            cursor.execute("SELECT id, store_name FROM stores WHERE is_active = TRUE")
            stores = cursor.fetchall()
            
            for store in stores:
                self.store_filter_combo.addItem(store['store_name'], store['id'])
                
            # تحديد متجر المستخدم الحالي كافتراضي
            user_store_id = self.user_info.get('store_id')
            if user_store_id:
                idx = self.store_filter_combo.findData(user_store_id)
                if idx >= 0:
                    self.store_filter_combo.setCurrentIndex(idx)
                    
        except Exception as e:
            print(f"خطأ: {e}")
    
    def add_product(self):
        """إضافة منتج جديد"""
        dialog = ProductDialog(self.db, None, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_all_data()
            self.data_changed.emit()

    def edit_product(self, product_id):
        """تعديل بيانات المنتج"""
        try:
            cursor = self.db.cursor
            cursor.execute("SELECT * FROM products WHERE id = %s", (product_id,))
            product = cursor.fetchone()
            
            if product:
                dialog = ProductDialog(self.db, product, self)
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    self.refresh_all_data()
                    self.data_changed.emit()
        except Exception as e:
            print(f"Error editing product: {e}")
    def refresh_all_data(self):
        """تحديث كافة البيانات في جميع التبويبات"""
        self.load_products()
        self.load_inventory()
        self.load_barcode_products()
        if hasattr(self, 'incoming_transfers_widget'):
            self.incoming_transfers_widget.load_transfers()

    def open_manual_edit(self, product_id, store_id):
        """فتح نافذة التعديل اليدوي"""
        dialog = InventoryDialog(self.db, product_id, store_id, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventory()
            self.data_changed.emit()
            
    def open_transfer(self, product_id, from_store_id, product_name):
        """فتح نافذة التحويل"""
        # نرسل from_store_id (المتجر المصدر)
        dialog = TransferDialog(self.db, product_id, from_store_id, self.user_info.get('id'), product_name, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventory()
            self.data_changed.emit()
    
    def filter_products(self):
        """Filter products using server-side query."""
        self.current_products_page = 1
        self.load_products()
    def filter_inventory(self):
        """Filter inventory using server-side query."""
        self.current_inventory_page = 1
        self.load_inventory()
    def next_products_page(self):
        """الانتقال للصفحة التالية"""
        total_products = self._products_total_count if self._products_total_count else len(self.all_products)
        total_pages = max(1, (total_products + self.products_per_page - 1) // self.products_per_page)
        if self.current_products_page < total_pages:
            self.current_products_page += 1
            self.load_products()
    
    def previous_products_page(self):
        """الانتقال للصفحة السابقة"""
        if self.current_products_page > 1:
            self.current_products_page -= 1
            self.load_products()
    
    def delete_product(self, product_id, product_name):
        """حذف منتج مع مراعاة البيانات المرتبطة"""
        try:
            cursor = self.db.cursor
            
            # 1. التحقق من وجود بيانات مرتبطة
            tables_to_check = {
                'invoice_items': 'فواتير بيع',
                'transfer_items': 'تحويلات مخزنية',
                'order_items': 'طلبات',
                'temporary_invoice_items': 'فواتير مؤقتة'
            }
            
            dependencies = []
            for table, label in tables_to_check.items():
                cursor.execute(f"SELECT COUNT(*) as count FROM {table} WHERE product_id = %s", (product_id,))
                result = cursor.fetchone()
                if result and result['count'] > 0:
                    dependencies.append(f"- {label}: {result['count']} سجل")
            
            if dependencies:
                # المنتج مرتبط ببيانات أخرى
                msg = f"لا يمكن حذف المنتج '{product_name}' مباشرة لأنه مرتبط بالبيانات التالية:\n\n"
                msg += "\n".join(dependencies)
                msg += "\n\nما الإجراء الذي تريد اتخاذه؟"
                
                # إنشاء صندوق حوار مخصص
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("تعذر الحذف المباشر")
                msg_box.setText(msg)
                msg_box.setIcon(QMessageBox.Icon.Warning)
                
                btn_deactivate = msg_box.addButton("تعطيل المنتج (أرشفة)", QMessageBox.ButtonRole.ActionRole)
                btn_force_delete = msg_box.addButton("حذف نهائي (مع السجلات)", QMessageBox.ButtonRole.DestructiveRole)
                btn_cancel = msg_box.addButton(QMessageBox.StandardButton.Cancel)
                
                msg_box.exec()
                
                if msg_box.clickedButton() == btn_deactivate:
                    # خيار التعطيل (Soft Delete)
                    cursor.execute("UPDATE products SET is_active = FALSE WHERE id = %s", (product_id,))
                    self.db.conn.commit()
                    QMessageBox.information(self, "نجح", "تم تعطيل المنتج وإخفاؤه من القوائم.")
                    self.load_products()
                    self.load_inventory()
                    
                elif msg_box.clickedButton() == btn_force_delete:
                    # خيار الحذف النهائي (Cascading Delete)
                    confirm = QMessageBox.warning(
                        self,
                        "تحذير شديد",
                        "هل أنت متأكد تماماً؟\n\nسيتم حذف جميع سجلات المبيع والتحويلات المرتبطة بهذا المنتج نهائياً.\nلا يمكن التراجع عن هذا الإجراء!",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    
                    if confirm == QMessageBox.StandardButton.Yes:
                        # حذف من جميع الجداول المرتبطة
                        tables_to_delete = [
                            'invoice_items', 'transfer_items', 'order_items', 
                            'temporary_invoice_items', 'price_history', 'product_inventory'
                        ]
                        
                        for table in tables_to_delete:
                            cursor.execute(f"DELETE FROM {table} WHERE product_id = %s", (product_id,))
                        
                        # أخيراً حذف المنتج
                        cursor.execute("DELETE FROM products WHERE id = %s", (product_id,))
                        self.db.conn.commit()
                        
                        QMessageBox.information(self, "نجح", "تم حذف المنتج وجميع سجلاته بنجاح.")
                        self.load_products()
                        self.load_inventory()
                        self.load_barcode_products()
                        self.data_changed.emit()
                        
            else:
                # لا توجد تبعيات، حذف عادي
                reply = QMessageBox.question(
                    self,
                    "تأكيد الحذف",
                    f"هل أنت متأكد من حذف المنتج: {product_name}؟\n\nلا توجد مبيعات مرتبطة به.",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    # حذف بيانات المخزون أولاً
                    cursor.execute("DELETE FROM product_inventory WHERE product_id = %s", (product_id,))
                    # ثم حذف المنتج
                    cursor.execute("DELETE FROM products WHERE id = %s", (product_id,))
                    self.db.conn.commit()
                    
                    QMessageBox.information(self, "نجح", "تم حذف المنتج بنجاح")
                    self.load_products()
                    self.load_inventory()
                    self.load_barcode_products()
                    self.data_changed.emit()

        except Exception as e:
            self.db.conn.rollback()
            QMessageBox.critical(self, "خطأ", f"فشل الإجراء: {str(e)}")
    
    def next_inventory_page(self):
        """الانتقال للصفحة التالية في المخزون"""
        total_inventory = self._inventory_total_count if self._inventory_total_count else len(self.all_inventory)
        total_pages = max(1, (total_inventory + self.inventory_per_page - 1) // self.inventory_per_page)
        if self.current_inventory_page < total_pages:
            self.current_inventory_page += 1
            self.load_inventory()
    
    def previous_inventory_page(self):
        """الانتقال للصفحة السابقة في المخزون"""
        if self.current_inventory_page > 1:
            self.current_inventory_page -= 1
            self.load_inventory()
            
            
    def import_from_excel(self):
        """استيراد سريع من Excel عبر دفعات (Batch Import)."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "اختر ملف Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return

        if not self.db or not self.db.cursor or not self.db.conn:
            QMessageBox.critical(self, "خطأ", "لا يوجد اتصال بقاعدة البيانات.")
            return

        def to_text(value):
            if value is None:
                return ""
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()

        def to_int(value, default=0):
            if value in (None, ""):
                return default
            try:
                if isinstance(value, bool):
                    return default
                return int(float(value))
            except (TypeError, ValueError):
                return default

        def to_float(value, default=0.0):
            if value in (None, ""):
                return default
            try:
                if isinstance(value, bool):
                    return default
                return float(value)
            except (TypeError, ValueError):
                return default

        wb = None
        progress = None
        started_at = perf_counter()

        try:
            cursor = self.db.cursor
            conn = self.db.conn

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            total_rows = max((ws.max_row or 1) - 1, 0)
            progress = QProgressDialog("جارٍ تجهيز الاستيراد...", "إلغاء", 0, max(total_rows, 1), self)
            progress.setWindowTitle("استيراد المنتجات")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)

            cursor.execute("SELECT id FROM stores WHERE is_active = TRUE ORDER BY id")
            store_ids = [int(row['id']) for row in cursor.fetchall()]
            if not store_ids:
                raise Exception("لا توجد فروع نشطة لإضافة المخزون.")

            store_id = int(self.user_info.get('store_id') or store_ids[0])
            if store_id not in store_ids:
                store_ids.append(store_id)

            cursor.execute("SELECT id, name FROM suppliers")
            supplier_cache = {}
            for row in cursor.fetchall():
                supplier_name = str(row.get('name') or "").strip().lower()
                if supplier_name:
                    supplier_cache[supplier_name] = int(row['id'])

            cursor.execute("SELECT id FROM categories")
            category_ids = {int(row['id']) for row in cursor.fetchall() if row.get('id') is not None}
            if not category_ids:
                cursor.execute(
                    "INSERT INTO categories (category_name, description) VALUES (%s, %s)",
                    ("عام", "تصنيف افتراضي تم إنشاؤه تلقائياً")
                )
                default_category_id = int(cursor.lastrowid)
                category_ids.add(default_category_id)
            else:
                default_category_id = 1 if 1 in category_ids else min(category_ids)

            product_upsert_sql = """
                INSERT INTO products (
                    product_code, product_name, category_id, buy_price,
                    sell_price, supplier_id, unit, barcode, description, is_active
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, NULL, NULL, TRUE)
                ON DUPLICATE KEY UPDATE
                    product_name = VALUES(product_name),
                    category_id = VALUES(category_id),
                    buy_price = VALUES(buy_price),
                    sell_price = VALUES(sell_price),
                    supplier_id = VALUES(supplier_id),
                    unit = VALUES(unit),
                    is_active = TRUE
            """

            inventory_upsert_sql = """
                INSERT INTO product_inventory (product_id, store_id, quantity_in_stock)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    quantity_in_stock = quantity_in_stock + VALUES(quantity_in_stock)
            """

            def fetch_product_ids(codes):
                code_to_id = {}
                fetch_chunk_size = 600
                for i in range(0, len(codes), fetch_chunk_size):
                    chunk_codes = codes[i:i + fetch_chunk_size]
                    placeholders = ", ".join(["%s"] * len(chunk_codes))
                    query = f"SELECT id, product_code FROM products WHERE product_code IN ({placeholders})"
                    cursor.execute(query, tuple(chunk_codes))
                    for product_row in cursor.fetchall():
                        code_to_id[str(product_row['product_code'])] = int(product_row['id'])
                return code_to_id

            def flush_batch(batch_map):
                if not batch_map:
                    return 0, 0

                records = list(batch_map.values())
                product_rows = [
                    (
                        rec['code'], rec['name'], rec['category_id'], rec['buy_price'],
                        rec['sell_price'], rec['supplier_id'], rec['unit']
                    )
                    for rec in records
                ]
                cursor.executemany(product_upsert_sql, product_rows)

                codes = [rec['code'] for rec in records]
                code_to_id = fetch_product_ids(codes)

                inventory_rows = []
                resolved = 0
                unresolved = 0
                for rec in records:
                    product_id = code_to_id.get(rec['code'])
                    if not product_id:
                        unresolved += 1
                        continue

                    resolved += 1
                    qty = rec['initial_qty']
                    for sid in store_ids:
                        inventory_rows.append((product_id, sid, qty if sid == store_id else 0))

                if inventory_rows:
                    cursor.executemany(inventory_upsert_sql, inventory_rows)

                return resolved, unresolved

            batch_limit = 2000
            batch_rows = 0
            batch_map = {}

            processed_rows = 0
            valid_rows = 0
            upsert_ops = 0
            created_suppliers = 0
            skipped = 0
            errors = 0
            unique_codes = set()
            cancelled = False
            error_samples = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if progress.wasCanceled():
                    cancelled = True
                    break

                processed_rows += 1

                try:
                    code = to_text(row[0] if len(row) > 0 else None)
                    name = to_text(row[1] if len(row) > 1 else None)

                    if not code or not name:
                        skipped += 1
                        continue

                    raw_category = row[2] if len(row) > 2 else None
                    category_id = to_int(raw_category, default_category_id)
                    if category_id not in category_ids:
                        category_id = default_category_id

                    buy_price = to_float(row[3] if len(row) > 3 else None, 0.0)
                    sell_price = to_float(row[4] if len(row) > 4 else None, 0.0)
                    unit = to_text(row[5] if len(row) > 5 else None) or "قطعة"
                    supplier_name = to_text(row[6] if len(row) > 6 else None)
                    initial_qty = to_int(row[7] if len(row) > 7 else None, 0)

                    if not supplier_name:
                        raise ValueError("المورد مطلوب")

                    supplier_key = supplier_name.lower()
                    supplier_id = supplier_cache.get(supplier_key)
                    if supplier_id is None:
                        cursor.execute(
                            """
                            INSERT INTO suppliers
                            (name, opening_balance, current_balance, notes, is_active)
                            VALUES (%s, 0, 0, %s, TRUE)
                            """,
                            (supplier_name, "تم إنشاؤه تلقائياً أثناء استيراد Excel")
                        )
                        supplier_id = int(cursor.lastrowid)
                        supplier_cache[supplier_key] = supplier_id
                        created_suppliers += 1

                    existing = batch_map.get(code)
                    if existing:
                        existing['initial_qty'] += initial_qty
                        existing['name'] = name
                        existing['category_id'] = category_id
                        existing['buy_price'] = buy_price
                        existing['sell_price'] = sell_price
                        existing['supplier_id'] = supplier_id
                        existing['unit'] = unit
                    else:
                        batch_map[code] = {
                            'code': code,
                            'name': name,
                            'category_id': category_id,
                            'buy_price': buy_price,
                            'sell_price': sell_price,
                            'supplier_id': supplier_id,
                            'unit': unit,
                            'initial_qty': initial_qty,
                        }

                    unique_codes.add(code)
                    valid_rows += 1
                    batch_rows += 1

                    if batch_rows >= batch_limit:
                        upserted_now, unresolved_now = flush_batch(batch_map)
                        upsert_ops += upserted_now
                        errors += unresolved_now
                        conn.commit()
                        batch_map.clear()
                        batch_rows = 0

                except Exception as row_err:
                    errors += 1
                    if len(error_samples) < 5:
                        error_samples.append(f"صف {processed_rows + 1}: {row_err}")

                if processed_rows % 200 == 0 or processed_rows == total_rows:
                    progress.setValue(min(processed_rows, max(total_rows, 1)))
                    progress.setLabelText(
                        f"جارٍ الاستيراد... {processed_rows}/{total_rows} | صالح: {valid_rows} | أكواد فريدة: {len(unique_codes)}"
                    )
                    QApplication.processEvents()

            if batch_map:
                upserted_now, unresolved_now = flush_batch(batch_map)
                upsert_ops += upserted_now
                errors += unresolved_now
                conn.commit()

            elapsed = perf_counter() - started_at
            progress.setValue(max(total_rows, 1))

            self.refresh_all_data()
            self.data_changed.emit()

            summary_lines = [
                f"الصفوف المقروءة: {processed_rows}",
                f"الصفوف الصالحة: {valid_rows}",
                f"أكواد المنتجات الفريدة: {len(unique_codes)}",
                f"عمليات الحفظ (Upsert): {upsert_ops}",
                f"موردون جدد: {created_suppliers}",
                f"الصفوف المتخطاة (بدون كود/اسم): {skipped}",
                f"الأخطاء: {errors}",
                f"المدة: {elapsed:.1f} ثانية",
            ]

            if error_samples:
                summary_lines.append("")
                summary_lines.append("أول الأخطاء:")
                summary_lines.extend(error_samples)

            if cancelled:
                QMessageBox.warning(
                    self,
                    "تم إيقاف الاستيراد",
                    "تم إيقاف العملية بواسطة المستخدم.\n\n" + "\n".join(summary_lines)
                )
            else:
                QMessageBox.information(
                    self,
                    "نجاح",
                    "تم الاستيراد بنجاح.\n\n" + "\n".join(summary_lines)
                )

        except Exception as e:
            try:
                self.db.conn.rollback()
            except Exception:
                pass
            QMessageBox.critical(self, "خطأ", f"خطأ في الاستيراد: {str(e)}")
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
            if progress:
                progress.close()
    
    def fix_zero_costs_dialog(self):
        """نافذة لتأكيد إصلاح التكاليف الصفرية"""
        msg = QMessageBox(self)
        msg.setWindowTitle("تصحيح التكاليف الصفرية")
        msg.setIcon(QMessageBox.Icon.Question)
        msg.setText("هل تريد تحديث سعر الشراء للمنتجات التي ليس لها تكلفة مسجلة؟")
        msg.setInformativeText("سيتم حساب التكلفة تلقائياً على أساس هامش ربح 25% من سعر البيع.\n(مثلاً: منتج يباع بـ 100 سيتم تسجيل تكلفته بـ 75)")
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        msg.setDefaultButton(QMessageBox.StandardButton.Cancel)
        
        if msg.exec() == QMessageBox.StandardButton.Yes:
            count = self.db.fix_zero_costs(margin_percent=25.0)
            if count > 0:
                QMessageBox.information(self, "تم", f"تم تحديث تكلفة {count} منتج بنجاح.")
                self.refresh_all_data()
                self.data_changed.emit()
            else:
                QMessageBox.information(self, "معلومة", "لم يتم العثور على منتجات بحاجة لتحديث (أو حدث خطأ).")

class ProductDialog(QDialog):
    """نافذة إضافة/تعديل منتج"""
    
    def __init__(self, db, product=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.product = product
        self.parent_page = parent # تعيين الصفحة الأب لتفادي خطأ AttributeError
        self.setWindowTitle("منتج جديد" if not product else "تعديل المنتج")
        self.resize(500, 600) # ضبط حجم النافذة بشكل مناسب
        self.init_ui()
    
    def init_ui(self):
        """إنشاء واجهة النافذة مع دعم التمرير"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # حاوية التمرير
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        scroll_content = QWidget()
        scroll_content.setStyleSheet("background: transparent;")
        layout = QFormLayout(scroll_content)
        layout.setSpacing(15)
        
        # الرمز مع زر التوليد التلقائي
        code_layout = QHBoxLayout()
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("أدخل رمز المنتج أو ولده تلقائياً")
        self.code_input.setStyleSheet(INPUT_STYLE)
        
        self.gen_btn = QPushButton("✨ توليد تلقائي")
        self.gen_btn.setStyleSheet(f"background-color: {COLORS['info']}; color: white; padding: 5px 15px; border-radius: 6px;")
        self.gen_btn.clicked.connect(self.generate_auto_barcode)
        self.gen_btn.setAutoDefault(False)
        
        self.code_input.textChanged.connect(self.toggle_generate_button)
        
        code_layout.addWidget(self.code_input)
        code_layout.addWidget(self.gen_btn)
        layout.addRow("رمز المنتج:", code_layout)
        
        # الاسم
        self.name_input = QLineEdit()
        self.name_input.setStyleSheet(INPUT_STYLE)
        layout.addRow("اسم المنتج:", self.name_input)
        
        # الفئة
        self.category_combo = QComboBox()
        self.category_combo.setStyleSheet(INPUT_STYLE)
        self.load_categories()
        layout.addRow("الفئة:", self.category_combo)
        
        # المورد (إجباري)
        self.supplier_combo = QComboBox()
        self.supplier_combo.setStyleSheet(INPUT_STYLE)
        self.load_suppliers()
        layout.addRow("المورد (إجباري):", self.supplier_combo)
        
        # سعر الشراء
        self.buy_price_label = QLabel("سعر الشراء:")
        self.buy_price = QDoubleSpinBox()
        self.buy_price.setStyleSheet(INPUT_STYLE)
        self.buy_price.setMaximum(100000)
        layout.addRow(self.buy_price_label, self.buy_price)
        
        # إخفاء سعر الشراء لغير المخولين
        can_view = getattr(self.parent_page, 'can_view_buy_price', True) if hasattr(self, 'parent_page') else True
        if not can_view:
            self.buy_price_label.hide()
            self.buy_price.hide()
        
        # سعر البيع
        self.sell_price = QDoubleSpinBox()
        self.sell_price.setStyleSheet(INPUT_STYLE)
        self.sell_price.setMaximum(100000)
        layout.addRow("سعر البيع:", self.sell_price)
        
        # الوحدة
        self.unit_input = QLineEdit()
        self.unit_input.setStyleSheet(INPUT_STYLE)
        self.unit_input.setPlaceholderText("مثلاً: كيلو، قطعة، كرتونة...")
        layout.addRow("الوحدة:", self.unit_input)
        
        # الحد الأدنى
        self.min_quantity = QSpinBox()
        self.min_quantity.setStyleSheet(INPUT_STYLE)
        layout.addRow("الحد الأدنى:", self.min_quantity)
        
        # رصيد أول المدة
        self.initial_stock = QSpinBox()
        self.initial_stock.setStyleSheet(INPUT_STYLE)
        self.initial_stock.setMaximum(100000)
        self.initial_stock.setValue(0)
        if not self.product:
            layout.addRow("رصيد أول المدة (الكمية الحالية):", self.initial_stock)
        else:
            self.initial_stock.hide()
        
        if self.product:
            self.load_product_data()
        
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)
        
        # الأزرار (خارج منطقة التمرير لتبقى ثابتة)
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 حفظ البيانات")
        save_btn.setStyleSheet(get_button_style('success'))
        save_btn.setFixedHeight(40)
        save_btn.setDefault(True)
        save_btn.clicked.connect(self.save_product)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("✖ إلغاء")
        cancel_btn.setStyleSheet(get_button_style('outline'))
        cancel_btn.setFixedHeight(40)
        cancel_btn.setAutoDefault(False)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
        self.toggle_generate_button()
    
    def toggle_generate_button(self):
        """تعطيل زر التوليد التلقائي إذا كان حقل الكود يحتوي على نص"""
        self.gen_btn.setEnabled(not bool(self.code_input.text().strip()))

    def load_categories(self):
        """تحميل الفئات"""
        try:
            get_cats_query = "SELECT id, category_name FROM categories"
            self.db.cursor.execute(get_cats_query)
            categories = self.db.cursor.fetchall()
            
            for cat in categories:
                self.category_combo.addItem(cat['category_name'], cat['id'])
        except Exception as e:
            print(f"Error loading categories: {e}")

    def load_suppliers(self):
        """تحميل الموردين"""
        try:
            self.supplier_combo.addItem("--- اختر المورد ---", None)
            # جلب الموردين النشطين
            self.db.cursor.execute("SELECT id, name FROM suppliers WHERE is_active = TRUE ORDER BY name")
            suppliers = self.db.cursor.fetchall()
            
            for s in suppliers:
                self.supplier_combo.addItem(s['name'], s['id'])
        except Exception as e:
            print(f"Error loading suppliers: {e}")
    
    def load_product_data(self):
        """تحميل بيانات المنتج"""
        self.code_input.setText(self.product['product_code'])
        self.name_input.setText(self.product['product_name'])
        self.buy_price.setValue(self.product['buy_price'])
        self.sell_price.setValue(self.product['sell_price'])
        self.unit_input.setText(self.product['unit'] or "")
        self.min_quantity.setValue(self.product.get('min_quantity', 0))
        
        if self.product.get('supplier_id'):
            idx = self.supplier_combo.findData(self.product['supplier_id'])
            if idx >= 0:
                self.supplier_combo.setCurrentIndex(idx)

    def generate_auto_barcode(self):
        """توليد باركود فريد تلقائياً"""
        import time
        import random
        # استخدام الوقت الحالي مع رقم عشوائي لضمان التفرد
        timestamp = str(int(time.time()))[2:] # حذف أول رقمين لتقصير الطول
        rand_suffix = str(random.randint(100, 999))
        auto_code = timestamp + rand_suffix
        self.code_input.setText(auto_code)
    
    def save_product(self):
        """حفظ المنتج"""
        code = self.code_input.text().strip()
        name = self.name_input.text().strip()
        category_id = self.category_combo.currentData()
        supplier_id = self.supplier_combo.currentData()
        purchase = self.buy_price.value()
        sale = self.sell_price.value()
        unit = self.unit_input.text().strip()
        
        can_view = getattr(self.parent_page, 'can_view_buy_price', True) if hasattr(self, 'parent_page') else True
        if not can_view and self.product:
            purchase = self.product.get('buy_price', 0.0)

        if not all([code, name]) or supplier_id is None:
            QMessageBox.warning(self, "خطأ", "يرجى ملء جميع الحقول المطلوبة واختيار المورد")
            return
        
        try:
            cursor = self.db.cursor
            if self.product:
                # تحديث منتج موجود
                cursor.execute("""
                    UPDATE products SET product_code=%s, product_name=%s,
                    category_id=%s, supplier_id=%s, buy_price=%s, sell_price=%s, unit=%s
                    WHERE id=%s
                """, (code, name, category_id, supplier_id, purchase, sale, unit, self.product['id']))
                
                # تحديث الحد الأدنى للمخزون
                current_store_id = self.parent_page.user_info.get('store_id') or 1
                self.db.update_product_min_stock(self.product['id'], current_store_id, self.min_quantity.value())
            else:
                # إضافة منتج جديد
                initial_qty = self.initial_stock.value()
                current_store_id = self.parent_page.user_info.get('store_id') or 1
                
                product_id = self.db.add_product(
                    product_code=code,
                    product_name=name,
                    category_id=category_id,
                    buy_price=purchase,
                    sell_price=sale,
                    supplier_id=supplier_id,
                    unit=unit,
                    barcode=code,
                    initial_stock=initial_qty,
                    store_id=current_store_id
                )
                
                # تحديث الحد الأدنى للمخزون
                if product_id:
                    self.db.update_product_min_stock(product_id, current_store_id, self.min_quantity.value())
            
            self.db.conn.commit()
            QMessageBox.information(self, "نجاح", "تم الحفظ بنجاح")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"خطأ: {str(e)}")

class IncomingTransfersWidget(QWidget):
    """ويدجت عرض واستلام التحويلات الواردة"""
    def __init__(self, db, user_info, parent=None):
        super().__init__(parent)
        self.db = db
        self.user_info = user_info
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # زر تحديث
        refresh_btn = QPushButton("🔃 تحديث القائمة")
        refresh_btn.setStyleSheet(f"background-color: {COLORS['info']}; color: white; border-radius: 8px; padding: 8px;")
        refresh_btn.clicked.connect(self.load_transfers)
        layout.addWidget(refresh_btn)
        
        # الجدول
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "رقم التحويل", "من المخزن", "المنتج", "الكمية", "تاريخ التحويل", "إجراء"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setStyleSheet(TABLE_STYLE)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        self.load_transfers()
        
    def load_transfers(self):
        """تحميل التحويلات المعلقة للمتجر الحالي"""
        to_store_id = self.user_info.get('store_id')
        user_role = self.user_info.get('role_name', '')
        
        # فقط مدير الفرع أو النظام يشوف ويقبل
        if user_role not in ['admin', 'manager']:
            self.table.setRowCount(0)
            return
            
        transfers = self.db.get_pending_transfers(to_store_id)
        self.table.setRowCount(len(transfers))
        
        for row, t in enumerate(transfers):
            self.table.setItem(row, 0, QTableWidgetItem(t['transfer_number']))
            self.table.setItem(row, 1, QTableWidgetItem(t['from_store']))
            self.table.setItem(row, 2, QTableWidgetItem(t['product_name']))
            self.table.setItem(row, 3, QTableWidgetItem(str(t['quantity_sent'])))
            self.table.setItem(row, 4, QTableWidgetItem(str(t['transfer_date'])))
            
            # زر الاستلام
            btn = QPushButton("📥 استلام")
            btn.setStyleSheet(f"background-color: {COLORS['success']}; color: white; border-radius: 4px;")
            btn.clicked.connect(lambda ch, tid=t['id']: self.receive_item(tid))
            self.table.setCellWidget(row, 5, btn)
            
    def receive_item(self, transfer_id):
        """استلام بضاعة"""
        reply = QMessageBox.question(self, "تأكيد", "هل أنت متأكد من استلام هذه الكمية وإضافتها للمخزون؟",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            success, msg = self.db.receive_transfer(transfer_id, self.user_info.get('id'))
            if success:
                QMessageBox.information(self, "نجاح", msg)
                self.load_transfers()
            else:
                QMessageBox.warning(self, "خطأ", msg)

class TransferDialog(QDialog):
    """نافذة تحويل المخزون (ديناميكية)"""
    
    def __init__(self, db, product_id, from_store_id, user_id, product_name, parent=None):
        super().__init__(parent)
        self.db = db
        self.product_id = product_id
        self.from_store_id = from_store_id
        self.user_id = user_id
        self.product_name = product_name
        
        self.setWindowTitle("نقل مخزون لفرع آخر")
        self.setFixedSize(400, 300)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"نقل المنتج: {self.product_name}"))
        
        form = QFormLayout()
        
        # المتجر المستلم
        self.store_combo = QComboBox()
        self.load_dest_stores()
        form.addRow("إلى المخزن/الفرع:", self.store_combo)
        
        # الكمية
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 100000)
        self.load_source_qty() # لضبط الحد الأقصى
        form.addRow("الكمية:", self.qty_spin)
        
        self.notes = QLineEdit()
        form.addRow("ملاحظات:", self.notes)
        
        layout.addLayout(form)
        
        btn_layout = QHBoxLayout()
        save = QPushButton("إرسال")
        save.clicked.connect(self.save)
        cancel = QPushButton("إلغاء")
        cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(save)
        btn_layout.addWidget(cancel)
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        
    def load_dest_stores(self):
        # تحميل كل المتاجر ما عدا المصدر
        try:
            cursor = self.db.cursor
            cursor.execute("SELECT id, store_name FROM stores WHERE id != %s AND is_active=TRUE", (self.from_store_id,))
            stores = cursor.fetchall()
            for s in stores:
                self.store_combo.addItem(s['store_name'], s['id'])
        except (mysql.connector.Error, AttributeError, KeyError, TypeError):
            pass
        
    def load_source_qty(self):
        try:
            cursor = self.db.cursor
            cursor.execute("SELECT quantity_in_stock FROM product_inventory WHERE product_id=%s AND store_id=%s", 
                           (self.product_id, self.from_store_id))
            res = cursor.fetchone()
            max_qty = res['quantity_in_stock'] if res else 0
            self.qty_spin.setMaximum(max_qty)
            self.setWindowTitle(f"نقل مخزون (المتاح: {max_qty})")
        except (mysql.connector.Error, AttributeError, KeyError, TypeError):
            pass

    def save(self):
        to_store = self.store_combo.currentData()
        qty = self.qty_spin.value()
        if not to_store: return
        
        success, msg = self.db.transfer_stock(self.product_id, self.from_store_id, to_store, qty, self.user_id, self.notes.text())
        if success:
            QMessageBox.information(self, "نجاح", msg)
            self.accept()
        else:
            QMessageBox.warning(self, "خطأ", msg)

class InventoryDialog(QDialog):
    """نافذة تعديل كمية المخزون (تعديل مباشر)"""
    def __init__(self, db, product_id, store_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.product_id = product_id
        self.store_id = store_id
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("تعديل الجرد (يدوي)")
        layout = QVBoxLayout()
        self.qty = QSpinBox()
        self.qty.setRange(0, 100000)
        
        self.min_qty = QSpinBox()
        self.min_qty.setRange(0, 10000)
        
        # Load current
        try:
            cursor = self.db.cursor
            cursor.execute("SELECT quantity_in_stock, minimum_quantity FROM product_inventory WHERE product_id=%s AND store_id=%s", (self.product_id, self.store_id))
            r = cursor.fetchone()
            if r: 
                self.qty.setValue(r['quantity_in_stock'])
                self.min_qty.setValue(r['minimum_quantity'])
        except (mysql.connector.Error, AttributeError, KeyError, TypeError):
            pass
        
        layout.addWidget(QLabel("الكمية الفعلية في المخزن:"))
        layout.addWidget(self.qty)
        
        layout.addWidget(QLabel("حد الأمان (الحد الأدنى):"))
        layout.addWidget(self.min_qty)
        
        save = QPushButton("حفظ")
        save.clicked.connect(self.save)
        layout.addWidget(save)
        self.setLayout(layout)
        
    def save(self):
        try:
            # تحديث الكمية
            self.db.update_inventory(self.product_id, self.store_id, self.qty.value(), 'set')
            # تحديث حد الأمان
            self.db.update_product_min_stock(self.product_id, self.store_id, self.min_qty.value())
            
            QMessageBox.information(self, "نجاح", "تم التحديث بنجاح")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "خطأ", str(e))

class IncomingTransfersWidget(QWidget):
    """ويدجت عرض واستلام التحويلات الواردة"""
    def __init__(self, db, user_info, parent=None, auto_load=True):
        super().__init__(parent)
        self.db = db
        self.user_info = user_info
        self.auto_load = auto_load
        self._is_loaded = False
        self.current_store_id = None
        self.init_ui()
        
    def set_user(self, user_info):
        self.user_info = user_info
        
        # تحديث ظهور الفلتر حسب الصلاحية الجديدة
        role = str(self.user_info.get('role_name', '')).lower()
        if hasattr(self, 'filter_combo'):
            if 'admin' in role:
                self.filter_combo.setVisible(True)
                # Reset selection if needed, or keep it
            else:
                self.filter_combo.setVisible(False)

        if self.auto_load or self._is_loaded:
            self.ensure_loaded(force=True)

    def ensure_loaded(self, force=False):
        """Load incoming transfers only when needed."""
        if force or not self._is_loaded:
            self.load_stores_filter()
            self.load_transfers()
            self._is_loaded = True

    def init_ui(self):
        layout = QVBoxLayout()
        
        # العنوان
        header_layout = QHBoxLayout()
        title = QLabel("التحويلات الواردة / المعلقة")
        title.setStyleSheet(LABEL_STYLE_TITLE)
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # فلتر الفروع (للأدمن فقط)
        self.filter_combo = QComboBox()
        self.filter_combo.addItem("كافة الفروع", None)
        if self.auto_load:
            self.load_stores_filter()
        self.filter_combo.currentIndexChanged.connect(lambda: self.load_transfers())
        
        # سيتم ضبط ظهوره في load_transfers أو set_user، ولكن هنا كبداية
        role = str(self.user_info.get('role_name', '')).lower()
        self.filter_combo.setVisible('admin' in role)
        
        header_layout.addWidget(QLabel("تصفية:"))
        header_layout.addWidget(self.filter_combo)
        
        # Label to hide context if hidden
        # Actually filter_combo visibility handles it, but label "تصفية:" should also be hidden?
        # Simpler: just set layout visible? No
        
        layout.addLayout(header_layout)
        
        # معلومات التصفية
        self.lbl_info = QLabel("")
        self.lbl_info.setStyleSheet(f"color: {COLORS['accent']}; font-weight: bold;")
        layout.addWidget(self.lbl_info)
        
        # زر تحديث
        refresh_btn = QPushButton("🔃 تحديث القائمة")
        refresh_btn.setStyleSheet(f"background-color: {COLORS['info']}; color: white; border-radius: 8px; padding: 8px;")
        refresh_btn.clicked.connect(lambda: self.load_transfers())
        layout.addWidget(refresh_btn)
        
        # الجدول
        self.table = QTableWidget()
        self.table.setColumnCount(7) # Increased col count for to_store
        self.table.setHorizontalHeaderLabels([
            "رقم التحويل", "من المخزن", "إلى المخزن", "المنتج", "الكمية", "تاريخ التحويل", "إجراء"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # تخصيص عمود الإجراءات (الأخير) ليكون بحجم ثابت وعريض
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(6, 250)
        
        # زيادة ارتفاع الصفوف
        self.table.verticalHeader().setDefaultSectionSize(60)
        
        self.table.setStyleSheet(TABLE_STYLE)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        # Initial load logic handling
        if self.auto_load:
            self.load_transfers()
            self._is_loaded = True
        
    def load_stores_filter(self):
        try:
            self.filter_combo.blockSignals(True)
            self.filter_combo.clear()
            self.filter_combo.addItem("كافة الفروع", None)
            cursor = self.db.cursor
            cursor.execute("SELECT id, store_name FROM stores WHERE is_active = TRUE")
            stores = cursor.fetchall()
            for s in stores:
                self.filter_combo.addItem(s['store_name'], s['id'])
            self.filter_combo.blockSignals(False)
        except (mysql.connector.Error, AttributeError, KeyError, TypeError):
            if hasattr(self, 'filter_combo'):
                self.filter_combo.blockSignals(False)
            pass

    def load_transfers(self, target_store_id=None):
        """تحميل التحويلات المعلقة
        target_store_id argument is kept for compatibility but Admin logic uses combo box.
        """
        user_role = str(self.user_info.get('role_name', '')).lower()
        try:
            user_store_id = int(self.user_info.get('store_id', 0))
        except (TypeError, ValueError):
            user_store_id = 0
            
        final_store_id = None
        
        if 'admin' in user_role:
            # للأدمن، نأخذ القيمة من الكومبو بوكس مباشرة
            combo_data = self.filter_combo.currentData()
            final_store_id = combo_data # Can be None (All) or specific ID
            
            store_text = self.filter_combo.currentText()
            self.lbl_info.setText(f"عرض التحويلات: {store_text}")
            
        elif 'manager' in user_role:
            final_store_id = user_store_id
            self.lbl_info.setText("التحويلات الواردة لفرعك")
        
        # جلب البيانات (None يعني الكل)
        transfers = self.db.get_pending_transfers(final_store_id)
        self.table.setRowCount(len(transfers))
        
        if len(transfers) == 0:
            if final_store_id:
                 self.lbl_info.setText(f"{self.lbl_info.text()} - (لا توجد تحويلات معلقة)")
            else:
                 self.lbl_info.setText("لا توجد أي تحويلات معلقة في النظام")
        
        for row, t in enumerate(transfers):
            self.table.setItem(row, 0, QTableWidgetItem(t['transfer_number']))
            self.table.setItem(row, 1, QTableWidgetItem(t['from_store']))
            # Handle potentially missing 'to_store_name' if old DB schema, but expected
            to_store_name = t.get('to_store_name', str(t.get('to_store_id', ''))) 
            self.table.setItem(row, 2, QTableWidgetItem(to_store_name))
            self.table.setItem(row, 3, QTableWidgetItem(t['product_name']))
            self.table.setItem(row, 4, QTableWidgetItem(str(t['quantity_sent'])))
            self.table.setItem(row, 5, QTableWidgetItem(str(t['transfer_date'])))
            
            # زر الاستلام
            # الأدمن يملك كل الصلاحيات
            # المدير يملك صلاحية فقط لمتجره
            can_receive = False
            try:
                row_to_store_id = int(t.get('to_store_id'))
            except (TypeError, ValueError):
                row_to_store_id = None
            
            if 'admin' in user_role:
                can_receive = True
            elif 'manager' in user_role and user_store_id == row_to_store_id:
                can_receive = True
                
            if can_receive:
                btn = QPushButton("📥 استلام")
                # تحسين مظهر الزر وزيادة حجمه
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {COLORS['success']};
                        color: white;
                        border-radius: 6px;
                        padding: 5px 20px;
                        font-weight: bold;
                        font-size: 14px;
                        min-height: 35px;
                        min-width: 100px;
                    }}
                    QPushButton:hover {{
                        background-color: #059669;
                    }}
                """)
                btn.setCursor(Qt.CursorShape.PointingHandCursor)
                btn.clicked.connect(lambda ch, tid=t['id']: self.receive_item(tid))
                
                # وضع الزر في حاوية لتوسيطه
                container = QWidget()
                layout = QHBoxLayout(container)
                layout.setContentsMargins(2, 2, 2, 2)
                layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                layout.addWidget(btn)
                
                self.table.setCellWidget(row, 6, container)
            else:
                self.table.setItem(row, 6, QTableWidgetItem("-"))
            
    def receive_item(self, transfer_id):
        """استلام بضاعة"""
        reply = QMessageBox.question(self, "تأكيد", "هل أنت متأكد من استلام هذه الكمية وإضافتها للمخزون؟",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            success, msg = self.db.receive_transfer(transfer_id, self.user_info.get('id'))
            if success:
                QMessageBox.information(self, "نجاح", msg)
                self.load_transfers()
            else:
                QMessageBox.warning(self, "خطأ", msg)

class CategoryManagementWidget(QWidget):
    """ويدجت إدارة فئات المنتجات"""
    def __init__(self, db, user_info, parent=None, auto_load=True):
        super().__init__(parent)
        self.db = db
        self.user_info = user_info
        self.auto_load = auto_load
        self._is_loaded = False
        self.init_ui()
        if self.auto_load:
            self.ensure_loaded(force=True)
        
    def set_user(self, user_info):
        self.user_info = user_info
        if self.auto_load or self._is_loaded:
            self.ensure_loaded(force=True)

    def ensure_loaded(self, force=False):
        if force or not self._is_loaded:
            self.load_categories()
            self._is_loaded = True

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # حاوية العنوان والأزرار (تنسيق متميز)
        header_container = QWidget()
        header_container.setStyleSheet(f"""
            QWidget {{
                background-color: {COLORS['primary']};
                border-radius: 12px;
                border: 1px solid {COLORS['border']};
            }}
        """)
        header_layout = QHBoxLayout(header_container)
        header_layout.setContentsMargins(15, 15, 15, 15)
        
        title = QLabel("📁 إدارة فئات المنتجات")
        title.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {COLORS['text_primary']}; border: none;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        add_btn = QPushButton("➕ إضافة فئة جديدة")
        add_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['success']}; 
                color: white; 
                padding: 12px 20px; 
                border-radius: 8px; 
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
        """)
        add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        add_btn.clicked.connect(self.add_category_dialog)
        header_layout.addWidget(add_btn)
        
        layout.addWidget(header_container)
        
        # الجدول
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["ID", "اسم الفئة", "الإجراءات"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(2, 220)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(TABLE_STYLE)
        
        # زيادة عرض (ارتفاع) الصفوف
        self.table.verticalHeader().setDefaultSectionSize(65)
        
        layout.addWidget(self.table)
        self.setLayout(layout)

    def load_categories(self):
        categories = self.db.get_categories()
        self.table.setRowCount(len(categories))
        
        for row, cat in enumerate(categories):
            # ID
            id_item = QTableWidgetItem(str(cat['id']))
            id_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 0, id_item)
            
            # Name
            name_item = QTableWidgetItem(cat['category_name'])
            name_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            name_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self.table.setItem(row, 1, name_item)
            
            # أزرار الإجراءات
            actions = QWidget()
            act_layout = QHBoxLayout(actions)
            act_layout.setContentsMargins(5, 5, 5, 5)
            act_layout.setSpacing(10)
            act_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            edit_btn = QPushButton("✏️")
            edit_btn.setToolTip("تعديل")
            edit_btn.setFixedSize(45, 45)
            edit_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: white;
                    border-radius: 6px;
                    font-size: 18px;
                    font-weight: 900;
                }}
                QPushButton:hover {{ background-color: {COLORS['secondary']}; }}
            """)
            edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            edit_btn.clicked.connect(lambda ch, c=cat: self.edit_category_dialog(c))
            
            del_btn = QPushButton("🗑️")
            del_btn.setToolTip("حذف")
            del_btn.setFixedSize(45, 45)
            del_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['danger']};
                    color: white;
                    border-radius: 6px;
                    font-size: 18px;
                    font-weight: 900;
                }}
                QPushButton:hover {{ background-color: #B91C1C; }}
            """)
            del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            del_btn.clicked.connect(lambda ch, cid=cat['id']: self.delete_category(cid))
            
            act_layout.addWidget(edit_btn)
            act_layout.addWidget(del_btn)
            self.table.setCellWidget(row, 2, actions)

    def add_category_dialog(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "إضافة فئة", "اسم الفئة الجديدة:")
        if ok and name.strip():
            if self.db.add_category(name.strip()):
                QMessageBox.information(self, "نجاح", "تمت إضافة الفئة بنجاح")
                self.load_categories()
            else:
                QMessageBox.warning(self, "خطأ", "فشل في إضافة الفئة")

    def edit_category_dialog(self, category):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "تعديل فئة", "الاسم الجديد للفئة:", text=category['category_name'])
        if ok and name.strip() and name.strip() != category['category_name']:
            if self.db.update_category(category['id'], name.strip()):
                QMessageBox.information(self, "نجاح", "تم تحديث الفئة بنجاح")
                self.load_categories()
            else:
                QMessageBox.warning(self, "خطأ", "فشل في تحديث الفئة")

    def delete_category(self, category_id):
        reply = QMessageBox.question(self, "تأكيد الحذف", "هل أنت متأكد من حذف هذه الفئة؟",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            success, msg = self.db.delete_category(category_id)
            if success:
                QMessageBox.information(self, "نجاح", msg)
                self.load_categories()
            else:
                QMessageBox.warning(self, "خطأ", msg)

