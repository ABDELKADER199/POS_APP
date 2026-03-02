
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_formatted_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Import Template"

    # Define Headers
    headers = [
        "الكود (Barcode)", 
        "اسم المنتج (Name)", 
        "رقم الفئة (Category ID)", 
        "سعر الشراء (Buy Price)", 
        "سعر البيع (Sell Price)", 
        "الوحدة (Unit)", 
        "اسم المورد (Supplier Name)", 
        "رصيد أول المدة (Stock)"
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Add and style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Adjust column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

    # Add sample data for clarity
    sample_data = [
        ["1001", "منتج تجريبي 1", 1, 100.0, 150.0, "قطعة", "عصام عبدالدايم", 50],
        ["1002", "منتج تجريبي 2", 1, 50.0, 75.0, "كيلو", "عصام عبدالدايم", 20],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.alignment = center_alignment
            cell.border = thin_border

    # Save file
    file_path = "template_products_import.xlsx"
    wb.save(file_path)
    print(f"✅ Template created successfully: {file_path}")

if __name__ == "__main__":
    create_formatted_template()
